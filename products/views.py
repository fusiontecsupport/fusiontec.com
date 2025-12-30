from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, Http404, HttpResponseRedirect
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.utils import timezone
from django.conf import settings
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.core.mail.backends.smtp import EmailBackend
from functools import wraps
import json
import logging
from decimal import Decimal
from datetime import datetime, timedelta
import razorpay
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch, mm
from io import BytesIO
import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def _generate_dsc_proforma_pdf(
    *,
    customer_name: str,
    company_name: str,
    mobile: str,
    email: str,
    address: str,
    gst_number: str,
    reference_name: str,
    reference_contact: str,
    class_type: str,
    user_type: str,
    cert_type: str,
    validity: str,
    include_token: bool,
    include_installation: bool,
    quoted_price: str,
    submission_id: int,
):
    """Create a DSC Proforma Invoice PDF.

    Returns (buffer, filename)
    """
    buffer = BytesIO()
    width, height = A4
    margin = 18 * mm
    content_w = width - 2 * margin

    c = canvas.Canvas(buffer, pagesize=A4)

    # Header band
    c.setFillColorRGB(23/255, 63/255, 132/255)
    header_h = 90
    c.rect(0, height - header_h, width, header_h, fill=1, stroke=0)

    # Try to place logo on left (use fusiontec.jpg to match popup)
    logo_candidates = [
        os.path.join(settings.BASE_DIR, 'staticfiles', 'products', 'img', 'fusiontec.jpg'),
        os.path.join(settings.BASE_DIR, 'staticfiles', 'products', 'img', 'fusiontec.png'),
        os.path.join(settings.BASE_DIR, 'staticfiles', 'products', 'img', 'logo.png'),
    ]
    logo_path = next((p for p in logo_candidates if os.path.exists(p)), None)
    if logo_path and os.path.exists(logo_path):
        try:
            # Smaller logo, vertically centered in header
            logo_w = 16 * mm
            logo_h = 16 * mm
            logo_y = height - header_h + (header_h - logo_h) / 2
            c.drawImage(logo_path, margin, logo_y, width=logo_w, height=logo_h, mask='auto')
        except Exception:
            pass

    # Left header text
    c.setFillColorRGB(1, 1, 1)
    # Place text to the right of the logo, vertically aligned with it
    text_x = margin + (16 * mm) + (4 * mm)
    c.setFont("Helvetica-Bold", 12)
    # Nudge header text a bit lower to align with the logo visually
    small_y = logo_y + (16 * mm) - 7
    c.drawString(text_x, small_y, "Digital Signature Certificate (DSC)")
    c.setFont("Helvetica-Bold", 20)
    title_y = logo_y + (16 * mm) / 2 - 6
    c.drawString(text_x, title_y, "PROFORMA INVOICE")

    # Invoice meta on right
    inv_no = f"PI-{int(datetime.now().timestamp())}"
    inv_date = datetime.now().strftime("%d/%m/%Y")
    c.setFont("Helvetica", 10)
    c.drawRightString(width - margin, height - 48, f"Proforma Invoice No: {inv_no}")
    c.drawRightString(width - margin, height - 30, f"Date: {inv_date}")

    # Add a bit more breathing room below the header before seller/buyer blocks
    y = height - 120
    
    # Seller information
    c.setFont("Helvetica-Bold", 11)
    c.setFillColorRGB(0.22, 0.25, 0.32)
    c.drawString(margin, y, "FUSIONTEC SOFTWARE")
    y -= 15
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(margin, y, "Suite No 17A, 2nd Floor, Crescent Court")
    y -= 12
    c.drawString(margin, y, "Door No 963 EVR Periyar Salai, Flowers Road Post")
    y -= 12
    c.drawString(margin, y, "Chennai -600 084")
    y -= 12
    c.drawString(margin, y, "GSTIN/UIN: 33AABFF1764F1ZU")
    y -= 12
    c.drawString(margin, y, "State Name : Tamil Nadu, Code : 33")
    y -= 12
    c.drawString(margin, y, "Contact : 04446072076 / 98416 99963 / 98416 34345")
    y -= 12
    c.drawString(margin, y, "E-Mail : dsc@fusiontec.com")
    
    # Buyer information (right side)
    buyer_x = width - margin - 240  # Slightly wider for better text display
    y = height - 120
    c.setFont("Helvetica-Bold", 11)
    c.setFillColorRGB(0.22, 0.25, 0.32)
    c.drawString(buyer_x, y, "Bill To:")
    y -= 15
    c.setFillColorRGB(0.4, 0.4, 0.4)
    
    # Customer name (bold)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(buyer_x, y, customer_name)
    y -= 12
    
    # Company name if provided
    if company_name and company_name.strip():
        c.setFont("Helvetica", 9)
        c.drawString(buyer_x, y, company_name)
        y -= 12
    
    # Mobile number (user mobile, not reference)
    c.setFont("Helvetica", 9)
    c.drawString(buyer_x, y, f"Mobile: {mobile}")
    y -= 12
    
    # Email
    c.drawString(buyer_x, y, f"Email: {email}")
    y -= 12
    
    # Address if provided
    if address and address.strip():
        # Wrap address if very long
        addr_lines = []
        line = ""
        for word in address.split():
            test = (line + " " + word).strip()
            if c.stringWidth(test, "Helvetica", 9) > 200:
                addr_lines.append(line)
                line = word
            else:
                line = test
        if line:
            addr_lines.append(line)
        for al in addr_lines:
            c.drawString(buyer_x, y, f"Address: {al}" if al == addr_lines[0] else f"         {al}")
            y -= 12
    
    # GST number if provided
    if gst_number and gst_number.strip():
        c.drawString(buyer_x, y, f"GST: {gst_number}")
        y -= 12
    
    # Reference information
    c.drawString(buyer_x, y, f"Reference Name: {reference_name}")
    y -= 12
    c.drawString(buyer_x, y, f"Reference Mobile: {reference_contact}")
    # Optional add-ons section (Yes/No)
    y -= 18
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(0.22, 0.25, 0.32)
    c.drawString(buyer_x, y, "Optional add-ons")
    y -= 14
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    # Include Token?
    c.drawString(buyer_x, y, f"Include Token?  {'Yes' if include_token else 'No'}")
    y -= 12
    # Include Installation?
    c.drawString(buyer_x, y, f"Include Installation?  {'Yes' if include_installation else 'No'}")

    # Product details table (Tally-style with 5 columns, no Qty)
    y -= 30
    table_y = y
    header_h = 22
    row_h = 26
    
    # Column widths (removed Qty column)
    unit_w = 90
    gstpct_w = 50
    gstamt_w = 90
    linetotal_w = 110
    desc_w = content_w - (unit_w + gstpct_w + gstamt_w + linetotal_w)
    
    # Column x positions
    x_desc = margin
    x_unit = x_desc + desc_w
    x_gstpct = x_unit + unit_w
    x_gstamt = x_gstpct + gstpct_w
    x_total = x_gstamt + gstamt_w
    
    # Header band
    c.setFillColorRGB(23/255, 63/255, 132/255)
    c.rect(margin, table_y, content_w, header_h, fill=1, stroke=0)
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(1, 1, 1)
    # Center the text vertically in the header row
    header_text_y = table_y + (header_h / 2) + 3  # Center vertically with small offset for visual balance
    c.drawString(x_desc + 8, header_text_y, "Description")
    c.drawRightString(x_unit + unit_w - 8, header_text_y, "Unit Price")
    c.drawCentredString(x_gstpct + gstpct_w/2, header_text_y, "GST %")
    c.drawRightString(x_gstamt + gstamt_w - 8, header_text_y, "GST Amt")
    c.drawRightString(x_total + linetotal_w - 8, header_text_y, "Line Total")
    
    # Single product row
    table_y -= row_h
    c.setFillColorRGB(1, 1, 1)
    c.rect(margin, table_y, content_w, row_h, fill=1, stroke=1)
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.2, 0.2, 0.2)
    product_desc = f"DSC {class_type.upper()} - {user_type} - {cert_type} - {validity}"
    
    # Amounts
    quoted_price_val = float(quoted_price or 0)
    gst_rate = 18.0
    basic_amount = quoted_price_val / (1 + (gst_rate / 100))
    gst_amount = quoted_price_val - basic_amount
    
    # Handle long description with word wrapping
    desc_lines = []
    line = ""
    for word in product_desc.split():
        test = (line + " " + word).strip()
        if c.stringWidth(test, "Helvetica", 9) > desc_w - 16:  # 16 for padding
            if line:
                desc_lines.append(line)
                line = word
            else:
                desc_lines.append(word)
        else:
            line = test
    if line:
        desc_lines.append(line)
    
    # Draw description lines
    desc_y = table_y + 16
    for i, line in enumerate(desc_lines):
        c.drawString(x_desc + 8, desc_y - (i * 12), line)
    
    # Draw other columns
    c.drawRightString(x_unit + unit_w - 8, table_y + 16, f"{basic_amount:,.2f}")
    c.drawCentredString(x_gstpct + gstpct_w/2, table_y + 16, f"{int(gst_rate)}%")
    c.drawRightString(x_gstamt + gstamt_w - 8, table_y + 16, f"{gst_amount:,.2f}")
    c.drawRightString(x_total + linetotal_w - 8, table_y + 16, f"{quoted_price_val:,.2f}")
    
    # Vertical column lines (header + row)
    top_y = y
    bottom_y = table_y
    c.setStrokeColorRGB(0.85, 0.85, 0.85)
    c.setLineWidth(0.8)
    for x in [x_unit, x_gstpct, x_gstamt, x_total]:
        c.line(x, bottom_y, x, top_y + header_h)
    # Outer border
    c.setStrokeColorRGB(0.5, 0.5, 0.5)
    c.rect(margin, bottom_y, content_w, (top_y - bottom_y) + header_h, fill=0, stroke=1)
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(1)
    
    # Summary block on the right
    sum_label_x = width - margin - 150
    sum_value_x = width - margin
    sum_y = bottom_y - 40
    c.setFont("Helvetica", 10)
    c.setFillColorRGB(0.25, 0.25, 0.25)
    c.drawString(sum_label_x, sum_y, "Basic Total")
    c.drawRightString(sum_value_x, sum_y, f"{basic_amount:,.2f}")
    sum_y -= 18
    c.drawString(sum_label_x, sum_y, f"GST ({int(gst_rate)}%)")
    c.drawRightString(sum_value_x, sum_y, f"{gst_amount:,.2f}")
    sum_y -= 18
    c.drawString(sum_label_x, sum_y, "Subtotal")
    c.drawRightString(sum_value_x, sum_y, f"{quoted_price_val:,.2f}")
    sum_y -= 22
    c.setFont("Helvetica-Bold", 11)
    c.drawString(sum_label_x, sum_y, "Grand Total")
    c.drawRightString(sum_value_x, sum_y, f"INR {quoted_price_val:,.2f}")
    
    # Move y for terms block
    table_y = sum_y - 28

    # Terms & Conditions block (restored)
    table_y -= 28
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(0.22, 0.25, 0.32)
    c.drawString(margin, table_y, "Terms & Conditions:")
    table_y -= 15
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.35, 0.35, 0.35)
    terms = [
        "• This is a proforma invoice and not a tax invoice",
        "• Payment is required before processing",
        "• Prices are subject to change without notice",
        "• GST will be charged as applicable",
        "• Delivery timeline: 3-5 business days after payment confirmation",
        "• Support: 24/7 technical assistance included",
    ]
    for term in terms:
        c.drawString(margin, table_y, term)
        table_y -= 12
    # Contact and payment terms lines
    table_y -= 8
    c.setFont("Helvetica-Bold", 9)
    c.setFillColorRGB(0.22, 0.25, 0.32)
    c.drawString(margin, table_y, "Please Contact us for any clarification at 98416 99963")
    table_y -= 14
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(margin, table_y, "Terms: Payment due within 7 days. This PI is valid for 15 days from the issue date")

    # Footer - use same color as header with white text
    c.setFillColorRGB(23/255, 63/255, 132/255)
    c.rect(0, 0, width, 30, fill=1, stroke=0)
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(1, 1, 1)  # White text
    c.drawCentredString(width / 2, 20, "This is computer generated proforma invoice signature not required")
    
    c.save()
    buffer.seek(0)
    
    filename = f"DSC_Proforma_Invoice_{submission_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buffer, filename

@csrf_exempt
def dsc_download_pdf_api(request):
    """API endpoint to generate and download DSC proforma invoice PDF."""
    print(f"[DEBUG] dsc_download_pdf_api called with method: {request.method}")
    
    if request.method == 'POST':
        try:
            print(f"[DEBUG] Request body: {request.body}")
            data = json.loads(request.body)
            print(f"[DEBUG] Parsed data: {data}")
            
            # Generate PDF using the same function as email
            print(f"[DEBUG] Calling _generate_dsc_proforma_pdf...")
            print(f"[DEBUG] Data received: {data}")
            
            # Ensure all string fields are properly handled
            customer_name = str(data.get('customer_name', '')).strip()
            company_name = str(data.get('company_name', '')).strip()
            mobile = str(data.get('mobile', '')).strip()
            email = str(data.get('email', '')).strip()
            address = str(data.get('address', '')).strip()
            gst_number = str(data.get('gst_number', '')).strip()
            reference_name = str(data.get('reference_name', '')).strip()
            reference_contact = str(data.get('reference_contact', '')).strip()
            class_type = str(data.get('class_type', '')).strip()
            user_type = str(data.get('user_type', '')).strip()
            cert_type = str(data.get('cert_type', '')).strip()
            validity = str(data.get('validity', '')).strip()
            quoted_price = str(data.get('quoted_price', '0')).strip()
            submission_id = int(data.get('submission_id', 0))
            
            print(f"[DEBUG] Processed data:")
            print(f"  customer_name: '{customer_name}'")
            print(f"  company_name: '{company_name}'")
            print(f"  mobile: '{mobile}'")
            print(f"  email: '{email}'")
            print(f"  address: '{address}'")
            print(f"  gst_number: '{gst_number}'")
            print(f"  reference_name: '{reference_name}'")
            print(f"  reference_contact: '{reference_contact}'")
            print(f"  class_type: '{class_type}'")
            print(f"  user_type: '{user_type}'")
            print(f"  cert_type: '{cert_type}'")
            print(f"  validity: '{validity}'")
            print(f"  quoted_price: '{quoted_price}'")
            print(f"  submission_id: {submission_id}")
            
            pdf_buffer, pdf_filename = _generate_dsc_proforma_pdf(
                customer_name=customer_name,
                company_name=company_name,
                mobile=mobile,
                email=email,
                address=address,
                gst_number=gst_number,
                reference_name=reference_name,
                reference_contact=reference_contact,
                class_type=class_type,
                user_type=user_type,
                cert_type=cert_type,
                validity=validity,
                include_token=data.get('include_token', False),
                include_installation=data.get('include_installation', False),
                quoted_price=quoted_price,
                submission_id=submission_id,
            )
            
            print(f"[DEBUG] PDF generated successfully: {pdf_filename}")
            print(f"[DEBUG] Buffer size: {pdf_buffer.getbuffer().nbytes}")
            
            # Return PDF as response
            from django.http import HttpResponse
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{pdf_filename}"'
            print(f"[DEBUG] Response created, returning PDF")
            return response
            
        except Exception as e:
            print(f"[DEBUG] Error in dsc_download_pdf_api: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({'error': str(e)}, status=500)
    else:
        print(f"[DEBUG] Invalid method: {request.method}")
        return JsonResponse({'error': 'Invalid request method'}, status=405)

def _generate_proforma_pdf(
    *,
    customer_name: str,
    company_name: str,
    mobile: str,
    email: str,
    quantity: int,
    unit_price: float,
    gst_rate: float,
    token_amount: float,
    install_charges: float,
    include_token: bool,
    include_install: bool,
    product_header: str,
    full_product_name: str,
    submission_id: int,
):
    """Create a simple Proforma Invoice PDF mirroring the modal's data.

    Returns (buffer, filename)
    """
    buffer = BytesIO()
    width, height = A4
    margin = 18 * mm
    content_w = width - 2 * margin

    c = canvas.Canvas(buffer, pagesize=A4)

    # Header band
    c.setFillColorRGB(23/255, 63/255, 132/255)
    header_h = 90
    c.rect(0, height - header_h, width, header_h, fill=1, stroke=0)

    # Try to place logo on left (use fusiontec.jpg to match popup)
    logo_candidates = [
        os.path.join(settings.BASE_DIR, 'staticfiles', 'products', 'img', 'fusiontec.jpg'),
        os.path.join(settings.BASE_DIR, 'staticfiles', 'products', 'img', 'fusiontec.png'),
        os.path.join(settings.BASE_DIR, 'staticfiles', 'products', 'img', 'logo.png'),
    ]
    logo_path = next((p for p in logo_candidates if os.path.exists(p)), None)
    if logo_path and os.path.exists(logo_path):
        try:
            # Smaller logo, vertically centered in header
            logo_w = 16 * mm
            logo_h = 16 * mm
            logo_y = height - header_h + (header_h - logo_h) / 2
            c.drawImage(logo_path, margin, logo_y, width=logo_w, height=logo_h, mask='auto')
        except Exception:
            pass

    # Left header text
    c.setFillColorRGB(1, 1, 1)
    # Place text to the right of the logo, vertically aligned with it
    text_x = margin + (16 * mm) + (4 * mm)
    c.setFont("Helvetica-Bold", 12)
    # Nudge header text a bit lower to align with the logo visually
    small_y = logo_y + (16 * mm) - 7
    c.drawString(text_x, small_y, str(product_header or "Tally Product"))
    c.setFont("Helvetica-Bold", 20)
    title_y = logo_y + (16 * mm) / 2 - 6
    c.drawString(text_x, title_y, "PROFORMA INVOICE")

    # Invoice meta on right
    inv_no = f"PI-{int(datetime.now().timestamp())}"
    inv_date = datetime.now().strftime("%d/%m/%Y")
    c.setFont("Helvetica", 10)
    c.drawRightString(width - margin, height - 48, f"Proforma Invoice No: {inv_no}")
    c.drawRightString(width - margin, height - 30, f"Date: {inv_date}")

    # Add a bit more breathing room below the header before seller/buyer blocks
    y = height - 120
    
    # Remove logo - no longer needed
    # y position remains at height - 120 for seller text
    
    c.setFont("Helvetica-Bold", 11)
    c.setFillColorRGB(0.22, 0.25, 0.32)
    c.drawString(margin, y, "Seller")
    c.drawString(width/2, y, "Buyer")
    y -= 16
    c.setFont("Helvetica", 10)
    c.setFillColorRGB(0, 0, 0)
    # Seller
    seller_lines = [
        "FUSIONTEC SOFTWARE",
        "Suite No 17A, 2nd Floor, Crescent Court",
        "Door No 963 EVR Periyar Salai, Flowers Road Post",
        "Chennai -600 084",
        "GSTIN/UIN: 33AABFF1764F1ZU",
        "State Name : Tamil Nadu, Code : 33",
        "Contact : 04446072076 / 98416 99963 / 91766 99599",
        "E-Mail : tally@fusiontec.com"
    ]
    for i, line in enumerate(seller_lines):
        if i == 0:  # First line - FUSIONTEC SOFTWARE
            c.setFont("Helvetica-Bold", 10)
        else:  # Other lines
            c.setFont("Helvetica", 10)
        c.drawString(margin, y - i*12, line)  # Reduced line spacing from 14 to 12
    
    # Buyer
    buyer_lines = [customer_name]
    if company_name:
        buyer_lines.append(company_name)
    buyer_lines += [mobile, email]
    for i, line in enumerate(buyer_lines):
        c.drawString(width/2, y - i*12, str(line))  # Reduced line spacing from 14 to 12

    # Add more spacing between seller/buyer section and table
    y = y - max(len(seller_lines), len(buyer_lines)) * 12 - 30

    # Check if we need a page break due to long seller details
    if y < 150:  # If too close to bottom, add new page
        c.showPage()
        y = height - 120  # Reset y position for new page

    # Compute pricing
    qty = int(quantity or 1)
    unit = float(unit_price or 0)
    token_val = float(token_amount or 0) if include_token else 0.0
    install_val = float(install_charges or 0) if include_install else 0.0

    line_basic = unit * qty
    gst_pct_value = float(gst_rate or 0)
    line_gst = (line_basic * gst_pct_value) / 100.0
    line_total = line_basic + line_gst
    sub_total = line_total
    grand_total = sub_total + token_val + install_val

    def fmt(n):
        return f"{float(n):,.2f}"

    # Items table
    # Derive gst percentage string using computed values if needed
    gst_pct_display = f"{round((line_gst/line_basic)*100, 2) if line_basic else 0}%"
    data = [["Description", "Qty", "Unit Price", "GST %", "GST Amt", "Line Total"]]
    # Wrap long description to avoid overlapping into numeric columns
    desc_style = ParagraphStyle(name="Desc", fontName="Helvetica", fontSize=9, leading=11)
    desc_para = Paragraph(str(full_product_name), desc_style)
    data.append([desc_para, str(qty), fmt(unit), gst_pct_display, fmt(line_gst), fmt(line_total)])
    if include_token and token_val > 0:
        data.append(["Token Charges", "1", fmt(token_val), "-", "-", fmt(token_val)])
    if include_install and install_val > 0:
        data.append(["Installation", "1", fmt(install_val), "-", "-", fmt(install_val)])

    # Column widths sized to content width so nothing overflows
    # Balanced widths with ~2% slack; prioritize Unit Price and Line Total
    col_desc = content_w * 0.45
    col_qty  = content_w * 0.06
    col_unit = content_w * 0.15
    col_pct  = content_w * 0.08
    col_gst  = content_w * 0.11
    col_line = content_w * 0.13
    table = Table(data, colWidths=[col_desc, col_qty, col_unit, col_pct, col_gst, col_line])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.Color(23/255, 63/255, 132/255)),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("VALIGN", (0,0), (-1,0), "MIDDLE"),  # Center header row vertically
        ("ALIGN", (0,0), (-1,0), "CENTER"),   # Center header text horizontally
        ("VALIGN", (0,1), (-1,-1), "TOP"),    # Data rows top-aligned
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),   # Data columns right-aligned
        ("ALIGN", (0,1), (0,-1), "LEFT"),     # Description column left-aligned
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("FONTSIZE", (0,1), (-1,-1), 9),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))

    # Draw table
    w, h = table.wrapOn(c, content_w, height)
    table.drawOn(c, margin, y - h)
    y = y - h - 12

    # Totals table (right aligned)
    totals = [["Basic Total", fmt(line_basic)], [f"GST ({gst_pct_display})", fmt(line_gst)], ["Subtotal", fmt(sub_total)]]
    if include_token and token_val > 0:
        totals.append(["Token Amount", fmt(token_val)])
    if include_install and install_val > 0:
        totals.append(["Installation", fmt(install_val)])
    totals.append(["Grand Total", f"INR {fmt(grand_total)}"])

    t = Table(totals, colWidths=[80*mm, 40*mm])
    t.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "RIGHT"),
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("LINEABOVE", (0,-1), (-1,-1), 0.7, colors.black),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
    ]))

    wt, ht = t.wrapOn(c, 120*mm, height)
    t.drawOn(c, margin + (content_w - wt), y - ht)
    y = y - ht - 18

    # Contact clarification note
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.42, 0.45, 0.50)
    c.drawString(margin, y, "Please Contact us for any clarification at 98416 99963")
    y -= 14
    
    # Footer note
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.42, 0.45, 0.50)
    c.drawString(margin, y, "Terms: Payment due within 7 days. This PI is valid for 15 days from issue date.")
    
    # Add blue footer with signature text at bottom of page
    footer_height = 30
    footer_y = footer_height  # Position at bottom of page
    c.setFillColorRGB(23/255, 63/255, 132/255)  # Same blue as header
    c.rect(0, 0, width, footer_height, fill=1)
    
    # Footer text in white - centered
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(1, 1, 1)  # White text
    
    # Center the text horizontally with fine-tuning
    footer_text = "This is computer generated proforma invoice signature not required"
    text_width = c.stringWidth(footer_text, "Helvetica", 9)
    center_x = (width - text_width) / 2
    
    # Fine-tune centering if needed
    if text_width > 0:
        center_x = (width - text_width) / 2
    else:
        # Fallback to manual calculation if stringWidth fails
        estimated_char_width = 5.5
        estimated_text_width = len(footer_text) * estimated_char_width
        center_x = (width - estimated_text_width) / 2
    
    c.drawString(center_x, footer_y - 15, footer_text)

    c.showPage()
    c.save()

    filename_safe = str(product_header or "product").lower().replace(" ", "_")
    try:
        buffer.seek(0)
    except Exception:
        pass
    return buffer, f"{filename_safe}_PI.pdf"

from .models import (
    ProductMaster, ProductType, ProductItem, RateCardMaster, Customer, 
    QuoteSubmission, ContactSubmission, PaymentTransaction, 
    PaymentSettings, Applicant, QuoteRequest,
    ProductTypeMaster, ProductMasterV2, ProductSubMaster, RateCardEntry,
    ProductFormSubmission, DscPrice, DscSubmission, DscEnquiry, NewProduct,
)

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def test_api(request):
    """Simple test endpoint to check if the server is working."""
    return JsonResponse({'status': 'success', 'message': 'API is working'})

def admin_required(view_func):
    """Custom decorator to check if admin is logged in via session"""
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if request.session.get('admin_logged_in', False):
            return view_func(request, *args, **kwargs)
        else:
            return HttpResponseRedirect('/admin-login/')
    return _wrapped_view

# ============================================================================
# HOME PAGE & PRODUCT CATALOG
# ============================================================================

def index(request):
    """Home page with contact form and product showcase"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()

        if not all([name, email, phone, subject, message]):
            messages.error(request, 'Please fill out all required fields.')
            return redirect('index')

        # Save to database
        ContactSubmission.objects.create(
            name=name,
            email=email,
            phone=phone,
            subject=subject,
            message=message,
        )

        # Admin notification email
        email_html_content = render_to_string('products/contact_form_email.html', {
            'name': name,
            'email': email,
            'phone': phone,
            'subject': subject,
            'message': message,
        }, request=request)

        # User thank-you email
        user_thank_you_content = render_to_string('products/contact_form_thanks.html', {
            'name': name,
        }, request=request)

        try:
            # Send to admin
            email_msg = EmailMessage(
                subject=f"[Fusiontec Contact] - {subject}",
                body=email_html_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[settings.CONTACT_FORM_RECIPIENT],
            )
            email_msg.content_subtype = "html"
            
            # Embed logo inline using CID
            try:
                from django.contrib.staticfiles import finders
                logo_path = finders.find('products/img/fusiontec.png')
                if logo_path:
                    with open(logo_path, 'rb') as f:
                        email_msg.attach_inline('fusiontec_logo', f.read(), 'image/png')
            except Exception as _e:
                print(f"Inline logo attach failed (admin): {_e}")
            
            email_msg.send()

            # Send thank you to user
            user_email = EmailMessage(
                subject="Thanks for contacting Fusiontec!",
                body=user_thank_you_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            user_email.content_subtype = "html"
            
            # Embed logo inline using CID
            try:
                from django.contrib.staticfiles import finders
                logo_path = finders.find('products/img/fusiontec.png')
                if logo_path:
                    with open(logo_path, 'rb') as f:
                        user_email.attach_inline('fusiontec_logo', f.read(), 'image/png')
            except Exception as _e:
                print(f"Inline logo attach failed (user): {_e}")
            
            user_email.send()

            messages.success(request, 'Your message has been sent successfully.')
        except Exception as e:
            messages.error(request, f'Error sending message: {str(e)}')

    # Get active products for showcase using new database structure
    product_masters = ProductMaster.objects.filter(is_active=True).order_by('display_order')
    
    # Get product type masters for the new structure
    product_type_masters = ProductTypeMaster.objects.all().order_by('id')
    
    # Locate DSC type for quote button convenience
    dsc_type = ProductTypeMaster.objects.filter(prdt_desc__icontains='dsc').first()
    
    # Get payment information
    razorpay_infos = PaymentSettings.objects.filter(setting_type='razorpay', is_active=True)
    payment_infos = PaymentSettings.objects.filter(setting_type='qr_code', is_active=True)
    bank_infos = PaymentSettings.objects.filter(setting_type='bank_transfer', is_active=True)
    
    # Get new products for promotional popup
    new_products = NewProduct.objects.filter(is_active=True).order_by('display_order', '-created_at')
    
    context = {
        'product_masters': product_masters,
        'product_type_masters': product_type_masters,
        'dsc_type': dsc_type,
        'razorpay_infos': razorpay_infos,
        'payment_infos': payment_infos,
        'bank_infos': bank_infos,
        'new_products': new_products,
    }
    return render(request, 'products/index.html', context)

def _get_index_context(highlight_section=None):
    """Helper function to get the common context for index page variations"""
    product_masters = ProductMaster.objects.filter(is_active=True).order_by('display_order')
    product_type_masters = ProductTypeMaster.objects.all().order_by('id')
    dsc_type = ProductTypeMaster.objects.filter(prdt_desc__icontains='dsc').first()
    razorpay_infos = PaymentSettings.objects.filter(setting_type='razorpay', is_active=True)
    payment_infos = PaymentSettings.objects.filter(setting_type='qr_code', is_active=True)
    bank_infos = PaymentSettings.objects.filter(setting_type='bank_transfer', is_active=True)
    new_products = NewProduct.objects.filter(is_active=True).order_by('display_order', '-created_at')
    
    context = {
        'product_masters': product_masters,
        'product_type_masters': product_type_masters,
        'dsc_type': dsc_type,
        'razorpay_infos': razorpay_infos,
        'payment_infos': payment_infos,
        'bank_infos': bank_infos,
        'new_products': new_products,
    }
    
    if highlight_section:
        context['highlight_section'] = highlight_section
    
    return context

def about_page(request):
    """About page - renders home page with about section highlighted"""
    context = _get_index_context('about')
    return render(request, 'products/index.html', context)

def process_page(request):
    """Process page - renders home page with process section highlighted"""
    context = _get_index_context('process')
    return render(request, 'products/index.html', context)

def products_page(request):
    """Products page - renders home page with products section highlighted"""
    context = _get_index_context('products')
    return render(request, 'products/index.html', context)

def contact_page(request):
    """Contact page - renders home page with contact section highlighted"""
    context = _get_index_context('footer')
    return render(request, 'products/index.html', context)

def dsc_page(request):
    """DSC page - renders home page with DSC section highlighted"""
    context = _get_index_context('dsc')
    return render(request, 'products/index.html', context)

def net_banking_page(request):
    """Net Banking page - renders home page with net banking section highlighted"""
    context = _get_index_context('net_banking')
    return render(request, 'products/index.html', context)

def privacy_policy(request):
    """Privacy Policy page"""
    return render(request, 'products/privacy_policy.html')

def tally_prime_page(request):
    """Tally Prime product page"""
    return render(request, 'products/tally_prime.html')

def tally_prime_server_page(request):
    """Tally Prime Server product page"""
    return render(request, 'products/tally_prime_server.html')

def tally_prime_aws_page(request):
    """Tally Prime on AWS product page"""
    return render(request, 'products/tally_prime_aws.html')

def tally_software_service_page(request):
    """Tally Software Service product page"""
    return render(request, 'products/tally_software_service.html')

def tally_prime_4_page(request):
    """Tally 7.0 product page"""
    return render(request, 'products/tally_prime_4.html')

def emudhra_digital_signature_page(request):
    """eMudhra Digital Signature product page"""
    return render(request, 'products/emudhra_digital_signature.html')

def product_catalog(request):
    """Product catalog page showing all products"""
    # Redirect specific product searches to dedicated pages
    search_query = request.GET.get('search', '').strip()
    if search_query:
        # Redirect Tally Prime Server searches to dedicated page
        if search_query.lower() in ['tally prime server', 'tallyprime server', 'tally server']:
            return redirect('tally_prime_server')
        # Redirect Tally Prime on AWS searches to dedicated page
        if search_query.lower() in ['tally prime on aws', 'tallyprime on aws', 'tally aws', 'tally prime aws']:
            return redirect('tally_prime_aws')
        # Redirect Tally Software Service searches to dedicated page
        if search_query.lower() in ['tally software service', 'tally software services', 'tss', 'tally service']:
            return redirect('tally_software_service')
        # Redirect Tally 7.0 searches to dedicated page (also supports old Tally Prime 4.0 searches for backward compatibility)
        if search_query.lower() in ['tally 7.0', 'tally 7', 'tally prime 4.0', 'tallyprime 4.0', 'tally prime 4', 'tallyprime 4']:
            return redirect('tally_prime_4')
    
    product_masters = ProductMaster.objects.filter(is_active=True).order_by('display_order')
    
    # Get search query again (for non-redirected searches)
    search_query = request.GET.get('search', '')
    if search_query:
        product_items = ProductItem.objects.filter(
            Q(item_name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(features__icontains=search_query),
            is_active=True
        ).select_related('product_type', 'product_type__product_master')
    else:
        product_items = ProductItem.objects.filter(
            is_active=True
        ).select_related('product_type', 'product_type__product_master')
    
    # Pagination
    paginator = Paginator(product_items, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'product_masters': product_masters,
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'products/catalog.html', context)

def product_detail(request, product_id):
    """Individual product detail page"""
    product_item = get_object_or_404(ProductItem, id=product_id, is_active=True)
    
    # Get related products from same type
    related_items = ProductItem.objects.filter(
        product_type=product_item.product_type,
        is_active=True
    ).exclude(id=product_id)[:4]
    
    context = {
        'product_item': product_item,
        'related_items': related_items,
    }
    return render(request, 'products/product_detail.html', context)

def product_type_products(request, type_id):
    """Show all products under a specific product type"""
    product_type = get_object_or_404(ProductType, id=type_id, is_active=True)
    product_items = ProductItem.objects.filter(
        product_type=product_type,
        is_active=True
    ).order_by('display_order')
    
    context = {
        'product_type': product_type,
        'product_items': product_items,
    }
    return render(request, 'products/product_type.html', context)

def product_type_form(request, type_name):
    """Display product form for a specific product type with Sub Products and rates per sub product."""
    from django.utils.text import slugify
    
    # Try to find by slugified name first, then by exact name
    product_type = None
    for pt in ProductTypeMaster.objects.all():
        if slugify(pt.prdt_desc) == type_name:
            product_type = pt
            break
    
    if not product_type:
        product_type = get_object_or_404(ProductTypeMaster, prdt_desc__iexact=type_name)
    
    # All products for this type
    available_products = ProductMasterV2.objects.filter(product_type=product_type).order_by('id')
    
    # Fallback display object
    product_item = available_products.first() or type('DummyProduct', (), {
        'prdt_desc': product_type.prdt_desc,
        'id': 0,
    })()
    
    # Build nested data: each product → list of sub products with latest rate
    products_with_rates = []
    for product in available_products:
        subs_data = []
        for sp in product.sub_products.all().order_by('id'):
            latest_rate = RateCardEntry.objects.filter(sub_product=sp).order_by('-rate_date').first()
            if latest_rate is None:
                latest_rate = type('DefaultRate', (), {
                    'base_amt': 0,
                    'gst_percent': 0,
                    'token_amount': 0,
                    'installation_charge': 0,
                })()
            subs_data.append({'sub': sp, 'rate_card': latest_rate})
        products_with_rates.append({'product': product, 'subs': subs_data})
    
    context = {
        'product': product_item,
        'product_type': product_type,
        'available_products': available_products,
        'products_with_rates': products_with_rates,
    }
    return render(request, 'products/product_form.html', context)

def dsc_form(request):
    """DSC form page with pricing and purchase options."""
    return render(request, 'products/dsc_form.html')

def dsc_integrated_submission(request):
    """Handle integrated DSC form submission for both enquiry and purchase."""
    print(f"Request method: {request.method}")
    print(f"Request POST data: {request.POST}")
    print(f"Request FILES: {request.FILES}")
    
    if request.method == 'POST':
        try:
            # Get form data
            submission_type = request.POST.get('submission_type', 'enquiry')
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            mobile = request.POST.get('mobile', '').strip()
            company_name = request.POST.get('company_name', '').strip()
            gst_number = request.POST.get('gst_number', '').strip()
            address = request.POST.get('address', '').strip()
            reference_name = request.POST.get('reference_name', '').strip()
            reference_contact = request.POST.get('reference_contact', '').strip()
            
            # DSC options
            class_type = request.POST.get('class_type', '')
            user_type = request.POST.get('user_type', '')
            cert_type = request.POST.get('cert_type', '')
            validity = request.POST.get('validity', '')
            include_token = request.POST.get('include_token') == 'on'
            include_installation = request.POST.get('include_installation') == 'on'
            outside_india = request.POST.get('outside_india') == 'on'
            quoted_price = request.POST.get('quoted_price', '0')
            
            # Validate required fields
            if not all([name, email, mobile, class_type, user_type, cert_type, validity]):
                messages.error(request, 'Please fill out all required fields.')
                return JsonResponse({'success': False, 'message': 'Required fields missing'})
            
            # Enforce required uploads
            # Others is required for all scenarios
            if 'other_document' not in request.FILES or not request.FILES.get('other_document'):
                return JsonResponse({'success': False, 'message': 'Others document is required.'}, status=400)
            # IEC Certificate required only for DGFT class
            if (class_type or '').strip().lower() == 'dgft':
                if 'ifc_certificate' not in request.FILES or not request.FILES.get('ifc_certificate'):
                    return JsonResponse({'success': False, 'message': 'IEC Certificate is required for DGFT.'}, status=400)
            
            # Create or get customer
            customer, created = Customer.objects.get_or_create(
                email=email,
                defaults={
                    'name': name,
                    'company_name': company_name,
                    'mobile': mobile,
                }
            )
            
            # Update customer if exists
            if not created:
                customer.name = name
                customer.company_name = company_name
                customer.mobile = mobile
                customer.save()
            
            # Create DSC submission
            dsc_submission = DscSubmission.objects.create(
                name=name,
                email=email,
                mobile=mobile,
                address=address,
                company_name=company_name,
                gst_number=gst_number,
                reference_name=reference_name,
                reference_contact=reference_contact,
                class_type=class_type,
                user_type=user_type,
                cert_type=cert_type,
                validity=validity,
                include_token=include_token,
                include_installation=include_installation,
                outside_india=outside_india,
                quoted_price=quoted_price,
                submission_type=submission_type,
                status='submitted'
            )
            
            # Handle file uploads if present
            if request.FILES:
                if 'pan_copy' in request.FILES:
                    dsc_submission.pan_copy = request.FILES['pan_copy']
                if 'aadhar_copy' in request.FILES:
                    dsc_submission.aadhar_copy = request.FILES['aadhar_copy']
                if 'photo' in request.FILES:
                    dsc_submission.photo = request.FILES['photo']
                if 'gst_certificate' in request.FILES:
                    dsc_submission.gst_certificate = request.FILES['gst_certificate']
                if 'authorization_letter' in request.FILES:
                    dsc_submission.authorization_letter = request.FILES['authorization_letter']
                if 'company_pan' in request.FILES:
                    dsc_submission.company_pan = request.FILES['company_pan']
                if 'ifc_certificate' in request.FILES:
                    dsc_submission.ifc_certificate = request.FILES['ifc_certificate']
                if 'other_document' in request.FILES:
                    dsc_submission.other_document = request.FILES['other_document']
                dsc_submission.save()
            
            # Send admin notification email
            admin_email_content = render_to_string('products/dsc_form_email.html', {
                'submission': dsc_submission,
                'submission_type': submission_type.title(),
                'admin_email': 'dsc@fusiontec.com',
                'logo_cid': 'fusiontec_logo'
            }, request=request)

            # Generate proforma invoice PDF
            try:
                pdf_buffer, pdf_filename = _generate_dsc_proforma_pdf(
                    customer_name=name,
                    company_name=company_name or '',
                    mobile=mobile,
                    email=email,
                    address=address or '',
                    gst_number=gst_number or '',
                    reference_name=reference_name or '',
                    reference_contact=reference_contact or '',
                    class_type=class_type,
                    user_type=user_type,
                    cert_type=cert_type,
                    validity=validity,
                    include_token=include_token,
                    include_installation=include_installation,
                    quoted_price=quoted_price,
                    submission_id=dsc_submission.id,
                )
            except Exception as pdf_err:
                print(f"PDF generation failed: {pdf_err}")
                pdf_buffer = None
                pdf_filename = None

            # Send customer confirmation email with proforma invoice (no logo version)
            customer_email_content = render_to_string('products/dsc_form_thanks_no_logo.html', {
                'submission': dsc_submission,
                'submission_type': submission_type.title(),
                'customer_name': name,
            }, request=request)

            try:
                # Send emails synchronously (more reliable on shared hosting)
                # Use DSC mailbox credentials (same as quote flow)
                sender_email_cfg = 'dsc@fusiontec.com'
                app_password_cfg = 'pcjn sxte zvci tljs'
                sanitized_password = app_password_cfg.replace(' ', '')

                email_backend = EmailBackend(
                    host='smtp.gmail.com',
                    port=587,
                    username=sender_email_cfg,
                    password=sanitized_password,
                    use_tls=True,
                    fail_silently=False
                )

                # Admin email
                admin_email = EmailMessage(
                    subject=f"[Fusiontec DSC {submission_type.title()}] - {class_type.upper()} - {user_type} - {cert_type} - {validity} - {name}",
                    body=admin_email_content,
                    from_email=sender_email_cfg,
                    to=['dsc@fusiontec.com'],
                )
                admin_email.content_subtype = "html"
                admin_email.connection = email_backend
                try:
                    # Embed logo inline using CID
                    from django.contrib.staticfiles import finders
                    logo_path = finders.find('products/img/fusiontec.png')
                    if logo_path:
                        with open(logo_path, 'rb') as f:
                            admin_email.attach_inline('fusiontec_logo', f.read(), 'image/png')
                except Exception as _e:
                    print(f"Inline logo attach failed (admin): {_e}")
                try:
                    # Attach uploaded files if available
                    attachments = [
                        ('pan_copy', 'PAN Card'),
                        ('aadhar_copy', 'Aadhar Card'),
                        ('photo', 'Photo'),
                        ('gst_certificate', 'GST Certificate'),
                        ('authorization_letter', 'Authorization Letter'),
                        ('company_pan', 'Company PAN'),
                        ('ifc_certificate', 'IEC Certificate'),
                        ('other_document', 'Others'),
                    ]
                    for field, label in attachments:
                        file_field = getattr(dsc_submission, field, None)
                        if file_field and getattr(file_field, 'name', None):
                            try:
                                file_field.open('rb')
                                content = file_field.read()
                                admin_email.attach(file_field.name, content)
                            finally:
                                try:
                                    file_field.close()
                                except Exception:
                                    pass
                except Exception as attach_err:
                    print(f"Attachment processing failed: {attach_err}")
                admin_email.send()

                # Customer email
                customer_email = EmailMessage(
                    subject=f"DSC {submission_type.title()} Received - Fusiontec",
                    body=customer_email_content,
                    from_email=sender_email_cfg,
                    to=[email],
                    reply_to=['dsc@fusiontec.com']
                )
                customer_email.content_subtype = "html"
                customer_email.connection = email_backend
                
                # Attach proforma invoice PDF if generated
                if pdf_buffer and pdf_filename:
                    try:
                        pdf_bytes = pdf_buffer.getvalue()
                        if pdf_bytes and len(pdf_bytes) > 500:
                            customer_email.attach(pdf_filename, pdf_bytes, 'application/pdf')
                            print(f"[EMAIL] Attached PDF '{pdf_filename}' size={len(pdf_bytes)} bytes to customer email")
                        else:
                            print("[EMAIL][WARN] Generated PDF is empty or too small; skipping attachment")
                    except Exception as att_err:
                        print(f"[EMAIL][ERROR] Failed to attach PDF to customer email: {att_err}")
                
                customer_email.send()

                # Immediate success response after emails sent
                if submission_type == 'enquiry':
                    messages.success(request, 'Enquiry submitted successfully! We will contact you soon.')
                else:
                    messages.success(request, 'Purchase request submitted successfully! Please complete the payment to proceed.')

                return JsonResponse({
                    'success': True,
                    'message': f'{submission_type.title()} submitted successfully!',
                    'submission_id': dsc_submission.id,
                    'submission_type': submission_type
                })

            except Exception as e:
                # Even if async scheduling fails, do not block user
                messages.warning(request, f'{submission_type.title()} saved but there was an error scheduling confirmation emails: {str(e)}')
                return JsonResponse({
                    'success': True,
                    'message': f'{submission_type.title()} submitted successfully!',
                    'submission_id': dsc_submission.id,
                    'submission_type': submission_type
                })
                
        except Exception as e:
            messages.error(request, f'Error submitting {submission_type}: {str(e)}')
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})





def dsc_price_list_page(request):
    """DSC price list page with detailed pricing information."""
    return render(request, 'products/type_of_dsc.html')

def dsc_price_api(request):
    """Return price for a DSC combination.
    Strict match on all four fields; no fallbacks to partial matches.
    Params: class_type, user_type, cert_type, validity, outside (0/1)
    """
    class_type = (request.GET.get('class_type', '') or '').strip()
    user_type = (request.GET.get('user_type', '') or '').strip().lower()
    cert_type = (request.GET.get('cert_type', '') or '').strip().lower()
    validity = (request.GET.get('validity', '') or '').strip().lower()
    outside = request.GET.get('outside', '0') == '1'

    try:
        price_row = (
            DscPrice.objects
            .filter(
                class_type__iexact=class_type,
                user_type__iexact=user_type,
                cert_type__iexact=cert_type,
                validity__iexact=validity,
                is_active=True,
            )
            .first()
        )

        if price_row is None:
            return JsonResponse({'success': False, 'message': 'Price not configured for this selection'}, status=404)

        price = float(price_row.nett_amount)
        if outside:
            price += float(price_row.outside_india_surcharge or 0)
        return JsonResponse({
            'success': True,
            'price': price,
            'currency': 'INR',
            'components': {
                'dsc_charge': float(price_row.dsc_charge),
                'token_amount': float(price_row.token_amount),
                'installation_charge': float(price_row.installation_charge),
                'gst_percent': float(price_row.gst_percent),
            },
            'outside_surcharge': float(price_row.outside_india_surcharge or 0)
        })
    except Exception as e:
        # Handle cases like missing table before migrations
        return JsonResponse({'success': False, 'message': f'Pricing backend not ready: {str(e)}'}, status=404)

def dsc_options_api(request):
    """Return available DSC combinations for building dynamic buttons.
    Structure:
    {
      classes: ["class3", "dgft", ...],
      map: { class: { user_type: { cert_type: ["1y","2y",...] } } },
      defaults: { class_type, user_type, cert_type, validity }
    }
    """
    try:
        options = {}
        classes = []
        default_row = None
        for row in DscPrice.objects.filter(is_active=True).order_by('created_at'):
            if default_row is None:
                default_row = row
            # Normalize
            c = (row.class_type or '').strip()
            u = (row.user_type or '').strip().lower()
            t = (row.cert_type or '').strip().lower()
            v = (row.validity or '').strip().lower()
            if c not in options:
                options[c] = {}
                classes.append(c)
            if u not in options[c]:
                options[c][u] = {}
            if t not in options[c][u]:
                options[c][u][t] = []
            if v not in options[c][u][t]:
                options[c][u][t].append(v)
        if not classes:
            return JsonResponse({'success': True, 'classes': [], 'map': {}, 'defaults': {} })
        defaults = {
            'class_type': (default_row.class_type or '').strip(),
            'user_type': (default_row.user_type or '').strip().lower(),
            'cert_type': (default_row.cert_type or '').strip().lower(),
            'validity': (default_row.validity or '').strip().lower(),
        }
        return JsonResponse({'success': True, 'classes': classes, 'map': options, 'defaults': defaults})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
def dsc_enquiry_api(request):
    """Create a DSC enquiry from the landing page."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    try:
        payload = json.loads(request.body or '{}')
    except Exception:
        payload = request.POST
    try:
        name = (payload.get('name') or '').strip()
        email = (payload.get('email') or '').strip()
        mobile = (payload.get('mobile') or '').strip()
        address = (payload.get('address') or '').strip() or None
        company_name = (payload.get('company_name') or '').strip() or None
        gst_number = (payload.get('gst_number') or '').strip() or None
        class_type = payload.get('class_type')
        user_type = payload.get('user_type')
        cert_type = payload.get('cert_type')
        validity = payload.get('validity')
        include_token = bool(payload.get('include_token'))
        include_installation = bool(payload.get('include_installation'))
        outside_india = bool(payload.get('outside_india'))
        quoted_price = float(payload.get('quoted_price') or 0)

        if not (name and email and mobile and address):
            return JsonResponse({'success': False, 'message': 'Name, Email, Mobile and Address are required.'}, status=400)

        enquiry = DscEnquiry.objects.create(
            name=name,
            email=email,
            mobile=mobile,
            address=address,
            company_name=company_name,
            gst_number=gst_number,
            class_type=class_type or '',
            user_type=user_type or '',
            cert_type=cert_type or '',
            validity=validity or '',
            include_token=include_token,
            include_installation=include_installation,
            outside_india=outside_india,
            quoted_price=quoted_price,
        )

        # Send admin notification email using common template
        admin_email_content = render_to_string('products/generic_form_email.html', {
            'form_title': 'DSC Enquiry Request',
            'form_subtitle': 'A new Digital Signature Certificate enquiry has been submitted',
            'customer_info': {
                'name': name,
                'email': email,
                'mobile': mobile,
                'company': company_name or 'Not provided',
                'address': address or 'Not provided',
                'gst_number': gst_number or 'Not provided'
            },
            'product_info': {
                'product': 'Digital Signature Certificate',
                'class_type': class_type,
                'user_type': user_type,
                'cert_type': cert_type,
                'validity': validity,
                'include_token': 'Yes' if include_token else 'No',
                'include_installation': 'Yes' if include_installation else 'No',
                'outside_india': 'Yes' if outside_india else 'No'
            },
            'pricing_info': {
                'quoted_price': f'₹{quoted_price:,.2f}'
            },
            'form_name': 'DSC Enquiry Form',
            'submission_id': enquiry.id,
            'priority_high': True,
            'customer_email': email
        }, request=request)

        # Send customer thank you email using common template
        customer_thank_you_content = render_to_string('products/generic_form_thanks.html', {
            'thank_you_title': 'Thank You for Your DSC Enquiry!',
            'thank_you_subtitle': 'We\'ve received your Digital Signature Certificate enquiry',
            'customer_name': name,
            'form_type': 'DSC enquiry',
            'response_time': '24 hours',
            'submission_details': {
                'product': 'Digital Signature Certificate',
                'class_type': class_type,
                'user_type': user_type,
                'cert_type': cert_type,
                'validity': validity,
                'quoted_price': f'₹{quoted_price:,.2f}'
            },
            'next_steps': [
                'Our DSC experts will review your requirements',
                'We\'ll verify eligibility for your selected class type',
                'You\'ll receive a detailed quote with all charges',
                'We\'ll guide you through documentation requirements',
                'We\'ll discuss the application process and timeline'
            ],
            'important_info': 'DSC applications require proper documentation and verification. Our team will guide you through the entire process.',
            'support_email': 'dsc@fusiontec.com'
        }, request=request)

        try:
            # Use DSC mailbox credentials like quote flow
            sender_email_cfg = 'dsc@fusiontec.com'
            app_password_cfg = 'pcjn sxte zvci tljs'
            sanitized_password = app_password_cfg.replace(' ', '')

            email_backend = EmailBackend(
                host='smtp.gmail.com',
                port=587,
                username=sender_email_cfg,
                password=sanitized_password,
                use_tls=True,
                fail_silently=False
            )

            # Send to admin
            admin_email = EmailMessage(
                subject=f"[Fusiontec DSC Enquiry] - {class_type} {user_type} - {name}",
                body=admin_email_content,
                from_email=sender_email_cfg,
                to=['dsc@fusiontec.com'],
            )
            admin_email.content_subtype = "html"
            admin_email.connection = email_backend
            admin_email.send()

            # Send thank you to customer
            customer_email = EmailMessage(
                subject="DSC Enquiry Received - Fusiontec",
                body=customer_thank_you_content,
                from_email=sender_email_cfg,
                to=[email],
                reply_to=['dsc@fusiontec.com']
            )
            customer_email.content_subtype = "html"
            customer_email.connection = email_backend
            customer_email.send()

            print(f"DSC enquiry created successfully with ID: {enquiry.id} and emails sent")
        except Exception as email_error:
            print(f"Email sending failed: {email_error}")
            # Continue with enquiry even if email fails

        return JsonResponse({'success': True, 'id': enquiry.id})
    except Exception as exc:
        return JsonResponse({'success': False, 'message': str(exc)}, status=500)

@csrf_exempt
def dsc_documents_upload_api(request):
    """Accept and validate DSC document uploads with required rules per selection.

    Expects multipart/form-data with fields:
    - class_type: class3|dgft
    - user_type: individual|organization (ignored for dgft)
    - cert_type: signature|encryption|both (ignored for dgft)
    - reference_name, reference_mobile
    - Optional: name, email, mobile (to link/create customer)
    - Files: pan, aadhar, gst, photo, auth, orgpan, ifc
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)

    try:
        class_type = (request.POST.get('class_type') or '').strip()
        user_type = (request.POST.get('user_type') or '').strip()
        cert_type = (request.POST.get('cert_type') or '').strip()
        reference_name = (request.POST.get('reference_name') or '').strip()
        reference_mobile = (request.POST.get('reference_mobile') or '').strip()

        name = (request.POST.get('name') or '').strip()
        email = (request.POST.get('email') or '').strip()
        mobile = (request.POST.get('mobile') or '').strip()

        if not reference_name or not reference_mobile:
            return JsonResponse({'success': False, 'message': 'Reference Name and Reference Mobile Number are required.'}, status=400)

        # Normalize inputs to handle variants
        def norm(s):
            return (s or '').strip().lower().replace('organisation', 'organization').replace(' ', '')
        ct = norm(class_type)
        ut = norm(user_type)
        cert = norm(cert_type)
        # Map common variants
        if ct in ('class3', 'classiii', 'class_3', 'class-3'):
            ct = 'class3'
        if ct in ('dgft', 'd.g.f.t'):
            ct = 'dgft'
        if ut in ('ind', 'indi', 'individual'):
            ut = 'individual'
        if ut in ('org', 'orga', 'organization'):
            ut = 'organization'
        if cert in ('sig', 'sign', 'signature'):
            cert = 'signature'
        if cert in ('enc', 'encrypt', 'encryption'):
            cert = 'encryption'
        if cert in ('both', 'combo', 'combined'):
            cert = 'both'

        # Determine required docs
        def required_docs(ct_key, ut_key, cert_key):
            if ct_key == 'dgft':
                return ['pan','aadhar','gst','photo','auth','orgpan','ifc']
            if ct_key == 'class3':
                if ut_key == 'individual' and cert_key in ('signature','encryption'):
                    return ['pan','aadhar']
                if ut_key == 'individual' and cert_key == 'both':
                    return ['pan','aadhar','gst','photo','auth','orgpan']
                if ut_key == 'organization':
                    return ['pan','aadhar','gst','photo','auth','orgpan']
            return []

        needed = required_docs(ct, ut, cert)
        if not needed:
            return JsonResponse({
                'success': False,
                'message': 'Invalid selection for documents.',
                'debug': {
                    'class_type': class_type,
                    'user_type': user_type,
                    'cert_type': cert_type,
                    'normalized': {'class_type': ct, 'user_type': ut, 'cert_type': cert}
                }
            }, status=400)

        # Collect files
        files = {
            'pan': request.FILES.get('pan'),
            'aadhar': request.FILES.get('aadhar'),
            'gst': request.FILES.get('gst'),
            'photo': request.FILES.get('photo'),
            'auth': request.FILES.get('auth'),
            'orgpan': request.FILES.get('orgpan'),
            'ifc': request.FILES.get('ifc'),
        }

        # Validate presence
        missing = [k for k in needed if not files.get(k)]
        if missing:
            return JsonResponse({'success': False, 'message': f"Missing required files: {', '.join(missing)}"}, status=400)

        # Validate extensions
        def has_allowed(name, allowed_exts):
            n = (name or '').lower()
            return any(n.endswith(ext) for ext in allowed_exts)

        doc_exts = ['.pdf', '.doc', '.docx']
        photo_exts = ['.jpg', '.jpeg', '.pdf']
        labels = {'pan':'PAN Card','aadhar':'Aadhar Card','gst':'GST Certificate','photo':'Photo','auth':'Authorisation Letter','orgpan':'Organisation PAN','ifc':'IFC Certificate'}

        for key, f in files.items():
            if not f:
                continue
            allowed = photo_exts if key == 'photo' else doc_exts
            if not has_allowed(getattr(f, 'name', ''), allowed):
                return JsonResponse({'success': False, 'message': f"{labels.get(key, key)} has invalid type."}, status=400)

        # Find or create a customer if we have minimal identity
        customer = None
        try:
            if mobile:
                customer = Customer.objects.filter(mobile=mobile).order_by('-id').first()
            if not customer and (name and (email or mobile)):
                customer = Customer.objects.create(name=name, email=email or None, mobile=mobile or None)
        except Exception:
            customer = None

        # Store Applicant with files
        app_kwargs = {
            'reference': reference_name,
            'reference_contact': reference_mobile,
        }
        if customer:
            app_kwargs['customer'] = customer
        else:
            # Create a placeholder customer if none
            customer = Customer.objects.create(name=name or reference_name, email=email or None, mobile=mobile or reference_mobile)
            app_kwargs['customer'] = customer

        applicant = Applicant.objects.create(**app_kwargs)

        # Assign files to model fields and save
        updated = False
        if files['pan']:
            applicant.pan_copy = files['pan']; updated = True
        if files['aadhar']:
            applicant.aadhar_copy = files['aadhar']; updated = True
        if files['photo']:
            applicant.photo = files['photo']; updated = True
        if files['gst']:
            applicant.gst_certificate = files['gst']; updated = True
        if files['auth']:
            applicant.authorization_letter = files['auth']; updated = True
        if files['orgpan']:
            applicant.company_pan = files['orgpan']; updated = True
        if updated:
            applicant.save()

        # Email admin with summary
        try:
            admin_email_content = render_to_string('products/generic_form_email.html', {
                'form_title': 'DSC Documents Uploaded',
                'form_subtitle': 'A customer uploaded DSC documents',
                'customer_info': {
                    'name': customer.name,
                    'email': customer.email or 'Not provided',
                    'mobile': customer.mobile or 'Not provided',
                },
                'product_info': {
                    'product': 'Digital Signature Certificate',
                    'class_type': class_type,
                    'user_type': user_type or '-',
                    'cert_type': cert_type or '-',
                },
                'extra_info': {
                    'reference_name': reference_name,
                    'reference_mobile': reference_mobile,
                    'required_docs': ', '.join(needed)
                },
                'form_name': 'DSC Documents Upload',
                'submission_id': applicant.id,
                'priority_high': True,
                'customer_email': customer.email or ''
            })

            admin_email = EmailMessage(
                subject=f"[Fusiontec DSC Docs] - {class_type} {user_type} - {customer.name}",
                body=admin_email_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=['dsc@fusiontec.com']
            )
            admin_email.content_subtype = 'html'
            admin_email.send(fail_silently=True)
        except Exception:
            pass

        return JsonResponse({'success': True, 'applicant_id': applicant.id})

    except Exception as exc:
        return JsonResponse({'success': False, 'message': str(exc)}, status=500)

@csrf_exempt
def dsc_submission_api(request):
    """Create a DSC submission for actual purchase with payment."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        print(f"DSC submission API called. Request body: {request.body}")
        print(f"Request content type: {request.content_type}")
        
        # Try to parse JSON first
        try:
            payload = json.loads(request.body or '{}')
            print(f"JSON payload parsed successfully: {payload}")
        except json.JSONDecodeError as e:
            print(f"JSON decode error: {e}, falling back to POST data")
            # Fallback to POST data if JSON parsing fails
            payload = request.POST
            print(f"POST data: {payload}")
        
        name = (payload.get('name') or '').strip()
        email = (payload.get('email') or '').strip()
        mobile = (payload.get('mobile') or '').strip()
        address = (payload.get('address') or '').strip() or None
        company_name = (payload.get('company_name') or '').strip() or None
        gst_number = (payload.get('gst_number') or '').strip() or None
        class_type = payload.get('class_type')
        user_type = payload.get('user_type')
        cert_type = payload.get('cert_type')
        validity = payload.get('validity')
        include_token = bool(payload.get('include_token'))
        include_installation = bool(payload.get('include_installation'))
        outside_india = bool(payload.get('outside_india'))
        quoted_price = float(payload.get('quoted_price') or 0)

        print(f"Parsed data - name: {name}, email: {email}, mobile: {mobile}, price: {quoted_price}")

        if not (name and email and mobile):
            return JsonResponse({'success': False, 'message': 'Name, Email and Mobile are required.'}, status=400)

        submission = DscSubmission.objects.create(
            name=name,
            email=email,
            mobile=mobile,
            address=address,
            company_name=company_name,
            gst_number=gst_number,
            class_type=class_type or '',
            user_type=user_type or '',
            cert_type=cert_type or '',
            validity=validity or '',
            include_token=include_token,
            include_installation=include_installation,
            outside_india=outside_india,
            quoted_price=quoted_price,
        )
        
        # Send admin notification email using common template
        admin_email_content = render_to_string('products/generic_form_email.html', {
            'form_title': 'DSC Purchase Request',
            'form_subtitle': 'A new Digital Signature Certificate purchase request has been submitted',
            'customer_info': {
                'name': name,
                'email': email,
                'mobile': mobile,
                'company': company_name or 'Not provided',
                'address': address or 'Not provided',
                'gst_number': gst_number or 'Not provided'
            },
            'product_info': {
                'product': 'Digital Signature Certificate',
                'class_type': class_type,
                'user_type': user_type,
                'cert_type': cert_type,
                'validity': validity,
                'include_token': 'Yes' if include_token else 'No',
                'include_installation': 'Yes' if include_installation else 'No',
                'outside_india': 'Yes' if outside_india else 'No'
            },
            'pricing_info': {
                'quoted_price': f'₹{quoted_price:,.2f}'
            },
            'form_name': 'DSC Purchase Form',
            'submission_id': submission.id,
            'priority_high': True,
            'customer_email': email
        }, request=request)

        # Send customer thank you email using common template
        customer_thank_you_content = render_to_string('products/generic_form_thanks.html', {
            'thank_you_title': 'Thank You for Your DSC Purchase Request!',
            'thank_you_subtitle': 'We\'ve received your Digital Signature Certificate purchase request',
            'customer_name': name,
            'form_type': 'DSC purchase request',
            'response_time': '24 hours',
            'submission_details': {
                'product': 'Digital Signature Certificate',
                'class_type': class_type,
                'user_type': user_type,
                'cert_type': cert_type,
                'validity': validity,
                'quoted_price': f'₹{quoted_price:,.2f}'
            },
            'next_steps': [
                'Our DSC experts will review your purchase request',
                'We\'ll verify your eligibility for the selected class type',
                'You\'ll receive payment instructions and documentation requirements',
                'We\'ll guide you through the DSC issuance process',
                'Your DSC will be issued and delivered as per the selected options'
            ],
            'important_info': 'DSC applications typically require 2-3 business days for processing and issuance.',
            'support_email': 'dsc@fusiontec.com'
        }, request=request)

        try:
            # Use DSC mailbox credentials like quote flow
            sender_email_cfg = 'dsc@fusiontec.com'
            app_password_cfg = 'pcjn sxte zvci tljs'
            sanitized_password = app_password_cfg.replace(' ', '')

            email_backend = EmailBackend(
                host='smtp.gmail.com',
                port=587,
                username=sender_email_cfg,
                password=sanitized_password,
                use_tls=True,
                fail_silently=False
            )

            # Send to admin
            admin_email = EmailMessage(
                subject=f"[Fusiontec DSC Purchase] - {class_type} {user_type} - {name}",
                body=admin_email_content,
                from_email=sender_email_cfg,
                to=['dsc@fusiontec.com'],
            )
            admin_email.content_subtype = "html"
            admin_email.connection = email_backend
            admin_email.send()

            # Send thank you to customer
            customer_email = EmailMessage(
                subject="DSC Purchase Request Received - Fusiontec",
                body=customer_thank_you_content,
                from_email=sender_email_cfg,
                to=[email],
                reply_to=['dsc@fusiontec.com']
            )
            customer_email.content_subtype = "html"
            customer_email.connection = email_backend
            customer_email.send()

            print(f"DSC submission created successfully with ID: {submission.id} and emails sent")
        except Exception as email_error:
            print(f"Email sending failed: {email_error}")
            # Continue with submission even if email fails
        
        return JsonResponse({'success': True, 'id': submission.id, 'submission_id': submission.id})
    except Exception as exc:
        print(f"Error in dsc_submission_api: {exc}")
        return JsonResponse({'success': False, 'message': f'Server error: {str(exc)}'}, status=500)

@csrf_exempt
def test_dsc_payment_api(request):
    """Test endpoint for DSC payment API."""
    print("=== Test DSC Payment API Called ===")
    return JsonResponse({'status': 'success', 'message': 'Test API working'})

@csrf_exempt
def create_dsc_payment_order(request):
    """Create Razorpay order for DSC purchase."""
    print(f"=== DSC Payment Order API Called ===")
    print(f"Request method: {request.method}")
    print(f"Request headers: {dict(request.headers)}")
    print(f"Request body: {request.body}")
    
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)
    
    try:
        # Log the request for debugging
        print(f"Creating DSC payment order. Request body: {request.body}")
        print(f"Request body type: {type(request.body)}")
        print(f"Request body length: {len(request.body) if request.body else 0}")
        
        try:
            data = json.loads(request.body or '{}')
            print(f"JSON parsed successfully: {data}")
        except json.JSONDecodeError as json_error:
            print(f"JSON decode error: {json_error}")
            print(f"Raw body: {request.body}")
            return JsonResponse({'status': 'error', 'message': f'Invalid JSON: {str(json_error)}'}, status=400)
        
        submission_id = data.get('submission_id')
        amount = data.get('amount')
        
        print(f"Parsed data - submission_id: {submission_id}, amount: {amount}")
        
        if not submission_id or not amount:
            return JsonResponse({'status': 'error', 'message': 'Missing submission_id or amount'}, status=400)
        
        # Verify submission exists
        try:
            submission = DscSubmission.objects.get(id=submission_id)
            print(f"Found submission: {submission.name} - {submission.email}")
        except DscSubmission.DoesNotExist:
            print(f"Submission not found with ID: {submission_id}")
            # List all submissions for debugging
            all_submissions = DscSubmission.objects.all().values('id', 'name', 'email')[:10]
            print(f"Available submissions: {list(all_submissions)}")
            return JsonResponse({'status': 'error', 'message': 'Submission not found'}, status=404)
        
        # Convert amount to paise (Razorpay expects amount in paise)
        try:
            amount_paise = int(float(amount) * 100)
            print(f"Amount in paise: {amount_paise}")
            if amount_paise <= 0:
                return JsonResponse({'status': 'error', 'message': 'Invalid amount: must be greater than 0'}, status=400)
        except (ValueError, TypeError) as e:
            print(f"Amount conversion error: {e}")
            return JsonResponse({'status': 'error', 'message': f'Invalid amount format: {amount}'}, status=400)
        
        # Check Razorpay credentials
        print(f"Razorpay Key ID: {settings.RAZORPAY_KEY_ID}")
        print(f"Razorpay Key Secret: {'*' * len(settings.RAZORPAY_KEY_SECRET) if settings.RAZORPAY_KEY_SECRET else 'None'}")
        
        if not settings.RAZORPAY_KEY_ID or not settings.RAZORPAY_KEY_SECRET:
            print("Razorpay credentials are missing")
            return JsonResponse({
                'status': 'error', 
                'message': 'Razorpay credentials are not configured'
            }, status=500)
        
        # Create Razorpay client
        try:
            client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
            print("Razorpay client created successfully")
        except Exception as client_error:
            print(f"Failed to create Razorpay client: {client_error}")
            return JsonResponse({
                'status': 'error', 
                'message': f'Failed to create Razorpay client: {str(client_error)}'
            }, status=500)
        
        # Create order
        order_data = {
            'amount': amount_paise,
            'currency': 'INR',
            'receipt': f'dsc_submission_{submission_id}',
            'notes': {
                'submission_id': str(submission_id),
                'customer_name': submission.name,
                'customer_email': submission.email,
                'dsc_type': f"{submission.class_type}/{submission.user_type}/{submission.cert_type}/{submission.validity}"
            }
        }
        
        print(f"Creating Razorpay order with data: {order_data}")
        
        try:
            order = client.order.create(data=order_data)
            print(f"Razorpay order created successfully: {order['id']}")
        except Exception as order_error:
            print(f"Failed to create Razorpay order: {order_error}")
            print(f"Order error type: {type(order_error)}")
            print(f"Order error details: {str(order_error)}")
            return JsonResponse({
                'status': 'error', 
                'message': f'Failed to create Razorpay order: {str(order_error)}'
            }, status=500)
        
        response_data = {
            'status': 'success',
            'order_id': order['id'],
            'amount': amount_paise,
            'currency': 'INR',
            'key': settings.RAZORPAY_KEY_ID
        }
        
        print(f"Returning success response: {response_data}")
        return JsonResponse(response_data)
        
    except DscSubmission.DoesNotExist:
        print(f"Submission not found with ID: {submission_id}")
        return JsonResponse({'status': 'error', 'message': 'Submission not found'}, status=404)
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {e}")
        return JsonResponse({'status': 'error', 'message': f'Invalid JSON data: {str(e)}'}, status=400)
    except Exception as e:
        print(f"Unexpected error in create_dsc_payment_order: {e}")
        return JsonResponse({'status': 'error', 'message': f'Failed to create order: {str(e)}'}, status=500)

@csrf_exempt
def verify_dsc_payment(request):
    """Verify Razorpay payment for DSC submission."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

    try:
        data = json.loads(request.body or '{}')
        payment_id = data.get('razorpay_payment_id')
        order_id = data.get('razorpay_order_id')
        signature = data.get('razorpay_signature')
        submission_id = data.get('submission_id')

        if not (payment_id and order_id and signature and submission_id):
            return JsonResponse({'status': 'error', 'message': 'Missing required fields'}, status=400)

        # Create Razorpay client
        try:
            client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
        except Exception as client_error:
            return JsonResponse({
                'status': 'error', 
                'message': f'Failed to create Razorpay client: {str(client_error)}'
            }, status=500)

        # Verify signature
        try:
            client.utility.verify_payment_signature({
                'razorpay_order_id': order_id,
                'razorpay_payment_id': payment_id,
                'razorpay_signature': signature,
            })
        except razorpay.errors.SignatureVerificationError:
            return JsonResponse({'status': 'error', 'message': 'Invalid payment signature'}, status=400)
        except Exception as verify_error:
            return JsonResponse({
                'status': 'error', 
                'message': f'Payment signature verification failed: {str(verify_error)}'
            }, status=500)

        # Update submission
        submission = DscSubmission.objects.get(id=submission_id)
        submission.razorpay_payment_id = payment_id
        submission.razorpay_order_id = order_id
        submission.payment_status = 'completed'
        submission.status = 'payment_received'
        submission.save()



        return JsonResponse({'status': 'success'})
        
    except DscSubmission.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Submission not found'}, status=404)
    except json.JSONDecodeError as e:
        return JsonResponse({'status': 'error', 'message': f'Invalid JSON data: {str(e)}'}, status=400)

@admin_required
def custom_admin_dsc_enquiries(request):
    """List DSC enquiries in custom admin."""
    enquiries = DscEnquiry.objects.all().order_by('-created_at')
    paginator = Paginator(enquiries, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = { 'page_obj': page_obj }
    return render(request, 'products/admin/dsc_enquiries.html', context)

@admin_required
def custom_admin_dsc_submissions(request):
    """List DSC submissions (actual purchases) in custom admin."""
    submissions = DscSubmission.objects.all().order_by('-created_at')
    paginator = Paginator(submissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = { 'page_obj': page_obj }
    return render(request, 'products/admin/dsc_submissions.html', context)

@csrf_exempt
def save_product_submission(request):
    """Save product submission to database"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Create or get customer
            customer, created = Customer.objects.get_or_create(
                email=data.get('email'),
                defaults={
                    'name': data.get('customer_name'),
                    'company_name': data.get('company_name', ''),
                    'mobile': data.get('mobile'),
                    'has_gst': data.get('has_gst') == 'yes',
                    'gst_number': data.get('gst_number', ''),
                    'address': data.get('address', ''),
                    'state': data.get('state', ''),
                    'district': data.get('district', ''),
                    'pincode': data.get('pincode', ''),
                }
            )
            
            # Update customer if not created
            if not created:
                customer.name = data.get('customer_name')
                customer.company_name = data.get('company_name', '')
                customer.mobile = data.get('mobile')
                customer.has_gst = data.get('has_gst') == 'yes'
                customer.gst_number = data.get('gst_number', '')
                customer.address = data.get('address', '')
                customer.state = data.get('state', '')
                customer.district = data.get('district', '')
                customer.pincode = data.get('pincode', '')
                customer.save()
            
            # Get product and optional sub product
            product_item = ProductMasterV2.objects.get(id=data.get('product_id'))
            sub_product_id = data.get('sub_product_id')
            sub_product_obj = None
            if sub_product_id:
                try:
                    sub_product_obj = ProductSubMaster.objects.get(id=sub_product_id)
                except ProductSubMaster.DoesNotExist:
                    sub_product_obj = None
            
            # Calculate GST rates from amounts
            basic_amount = data.get('basic_amount', 0)
            cgst_amount = data.get('cgst', 0)
            sgst_amount = data.get('sgst', 0)
            
            # Calculate GST rates (avoid division by zero)
            cgst_rate = 0
            sgst_rate = 0
            if basic_amount > 0:
                cgst_rate = (cgst_amount / basic_amount) * 100
                sgst_rate = (sgst_amount / basic_amount) * 100
            
            # Save to the new ProductFormSubmission table
            form_submission = ProductFormSubmission.objects.create(
                customer_name=data.get('customer_name'),
                company_name=data.get('company_name', ''),
                mobile=data.get('mobile'),
                email=data.get('email'),
                has_gst=data.get('has_gst'),
                gst_number=data.get('gst_number', ''),
                address=data.get('address', ''),
                state=data.get('state', ''),
                district=data.get('district', ''),
                pincode=data.get('pincode', ''),
                product_id=product_item,
                quantity=data.get('quantity', 1),
                basic_amount=basic_amount,
                cgst_rate=cgst_rate,
                sgst_rate=sgst_rate,
                cgst_amount=cgst_amount,
                sgst_amount=sgst_amount,
                total_with_gst=data.get('total_amount', 0),
                token_amount=data.get('token_amount', 0),
                installation_charges=data.get('installing_charges', 0),
                grand_total=data.get('grand_total', 0),
                status='new'
            )
            
            # Send admin notification email (no logo version)
            admin_email_content = render_to_string('products/product_form_email_no_logo.html', {
                'customer_name': data.get('customer_name'),
                'company_name': data.get('company_name', ''),
                'email': data.get('email'),
                'mobile': data.get('mobile'),
                'has_gst': data.get('has_gst'),
                'gst_number': data.get('gst_number', ''),
                'address': data.get('address', ''),
                'state': data.get('state', ''),
                'district': data.get('district', ''),
                'pincode': data.get('pincode', ''),
                'product_name': f"{product_item.prdt_desc} — {sub_product_obj.subprdt_desc}" if sub_product_obj else product_item.prdt_desc,
                'quantity': data.get('quantity', 1),
                'basic_amount': basic_amount,
                'cgst_amount': cgst_amount,
                'sgst_amount': sgst_amount,
                'total_with_gst': data.get('total_amount', 0),
                'token_amount': data.get('token_amount', 0),
                'installation_charges': data.get('installing_charges', 0),
                'grand_total': data.get('grand_total', 0),
                'submission_id': form_submission.id,
            })

            # Send customer thank you email (no logo version)
            customer_thank_you_content = render_to_string('products/product_form_thanks_no_logo.html', {
                'customer_name': data.get('customer_name'),
                'product_name': product_item.prdt_desc,
                'quantity': data.get('quantity', 1),
                'submission_id': form_submission.id,
            })

            try:
                # Prepare emails (admin and customer)
                admin_email = EmailMessage(
                    subject=f"[Fusiontec Product Form] - {product_item.prdt_desc}{(' — ' + sub_product_obj.subprdt_desc) if sub_product_obj else ''} - {data.get('customer_name')}",
                    body=admin_email_content,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[settings.CONTACT_FORM_RECIPIENT],
                )
                admin_email.content_subtype = "html"

                customer_email = EmailMessage(
                    subject="Product Form Submission Received - Fusiontec",
                    body=customer_thank_you_content,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[data.get('email')],
                )
                customer_email.content_subtype = "html"

                # Generate single Proforma Invoice PDF and attach to both emails so they match the popup/download
                try:
                    # Determine GST % to mirror popup exactly
                    gst_pct_value = 0.0
                    try:
                        if sub_product_obj:
                            rc = RateCardEntry.objects.filter(sub_product=sub_product_obj).order_by('-rate_date').first()
                            if rc and rc.gst_percent is not None:
                                gst_pct_value = float(rc.gst_percent)
                    except Exception:
                        gst_pct_value = 0.0
                    if not gst_pct_value and basic_amount:
                        try:
                            total_with_gst_val = float(data.get('total_amount') or 0)
                            if total_with_gst_val and float(basic_amount) > 0:
                                gst_pct_value = ((total_with_gst_val - float(basic_amount)) / float(basic_amount)) * 100.0
                        except Exception:
                            pass

                    pdf_buffer, pdf_filename = _generate_proforma_pdf(
                        customer_name=data.get('customer_name'),
                        company_name=data.get('company_name') or '',
                        mobile=data.get('mobile'),
                        email=data.get('email'),
                        quantity=int(data.get('quantity') or 1),
                        unit_price=(float(basic_amount or 0) / max(int(data.get('quantity') or 1), 1)),
                        gst_rate=float(gst_pct_value or 0),
                        token_amount=float(data.get('token_amount') or 0),
                        install_charges=float(data.get('installing_charges') or 0),
                        include_token=float(data.get('token_amount') or 0) > 0,
                        include_install=float(data.get('installing_charges') or 0) > 0,
                        product_header=product_item.prdt_desc,
                        full_product_name=(f"{product_item.prdt_desc} — {sub_product_obj.subprdt_desc}" if sub_product_obj else product_item.prdt_desc),
                        submission_id=form_submission.id,
                    )

                    try:
                        pdf_bytes = pdf_buffer.getvalue()
                        print(f"[EMAIL] Attaching PDF '{pdf_filename}' size={len(pdf_bytes)} bytes for submission {form_submission.id}")
                        if pdf_bytes and len(pdf_bytes) > 500:
                            # Attach only to customer email as before
                            customer_email.attach(pdf_filename, pdf_bytes, 'application/pdf')
                        else:
                            print("[EMAIL][WARN] Generated PDF is empty or too small; skipping attachment")
                    except Exception as att_err:
                        print(f"[EMAIL][ERROR] Failed to attach PDF: {att_err}")
                except Exception as gen_err:
                    # Do not fail email if PDF generation fails
                    print(f"[EMAIL][ERROR] PDF generation failed: {gen_err}")

                # Send both emails
                admin_email.send()
                customer_email.send()

                return JsonResponse({
                    'status': 'success',
                    'message': 'Form submitted successfully! You will receive a confirmation email shortly.',
                    'form_submission_id': form_submission.id
                })
            except Exception as e:
                return JsonResponse({
                    'status': 'success',
                    'message': f'Form submitted successfully but there was an error sending confirmation emails: {str(e)}',
                    'form_submission_id': form_submission.id
                })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error saving submission: {str(e)}'
            }, status=500)
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

# ============================================================================
# PAYMENTS: Razorpay integration for product form modal
# ============================================================================

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def create_payment_order(request):
    """Create a Razorpay order for the given amount and return order details.

    Expects JSON body: { amount, customer_name, email, mobile, product_name, form_submission_id }
    Returns: { order_id, key, amount, currency }
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

    try:
        import json as _json
        payload = _json.loads(request.body or '{}')

        amount = float(payload.get('amount') or 0)
        customer_name = (payload.get('customer_name') or '').strip()
        email = (payload.get('email') or '').strip()
        mobile = (payload.get('mobile') or '').strip()
        product_name = (payload.get('product_name') or '').strip()
        form_submission_id = payload.get('form_submission_id')

        if amount <= 0:
            return JsonResponse({'status': 'error', 'message': 'Invalid amount'}, status=400)

        # Create order with Razorpay
        import razorpay as _razorpay
        client = _razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

        amount_paise = int(round(amount * 100))
        order = client.order.create({
            'amount': amount_paise,
            'currency': 'INR',
            'payment_capture': 1,
            'notes': {
                'form_submission_id': str(form_submission_id or ''),
                'customer_email': email,
                'customer_name': customer_name,
                'product_name': product_name,
            }
        })

        return JsonResponse({
            'status': 'success',
            'order_id': order.get('id'),
            'key': settings.RAZORPAY_KEY_ID,
            'amount': amount_paise,
            'currency': 'INR',
        })
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Failed to create order: {str(e)}'}, status=500)


@csrf_exempt
def verify_payment(request):
    """Verify Razorpay payment signature and record transaction.

    Expects JSON body with fields from Razorpay handler and customer context.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

    try:
        import json as _json
        from decimal import Decimal
        import razorpay as _razorpay

        data = _json.loads(request.body or '{}')
        payment_id = data.get('razorpay_payment_id')
        order_id = data.get('razorpay_order_id')
        signature = data.get('razorpay_signature')
        amount = Decimal(str(data.get('amount') or '0'))
        customer_name = (data.get('customer_name') or '').strip()
        email = (data.get('email') or '').strip()
        mobile = (data.get('mobile') or '').strip()
        form_submission_id = data.get('form_submission_id')

        if not (payment_id and order_id and signature):
            return JsonResponse({'status': 'error', 'message': 'Missing Razorpay fields'}, status=400)

        client = _razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

        # Verify signature
        client.utility.verify_payment_signature({
            'razorpay_order_id': order_id,
            'razorpay_payment_id': payment_id,
            'razorpay_signature': signature,
        })

        # Create or update customer record
        customer_obj, _ = Customer.objects.get_or_create(
            email=email,
            defaults={
                'name': customer_name or (email or mobile) or 'Customer',
                'company_name': '',
                'mobile': mobile or '',
                'has_gst': True,
            }
        )
        if customer_name:
            customer_obj.name = customer_name
        if mobile:
            customer_obj.mobile = mobile
        customer_obj.save()

        # Record transaction
        PaymentTransaction.objects.create(
            customer=customer_obj,
            quote=None,
            amount=amount,
            payment_method='razorpay',
            razorpay_payment_id=payment_id,
            razorpay_order_id=order_id,
            status='success',
        )

        # Optionally update form submission status
        try:
            if form_submission_id:
                pfs = ProductFormSubmission.objects.filter(id=form_submission_id).first()
                if pfs:
                    pfs.status = 'approved'
                    pfs.admin_notes = (pfs.admin_notes or '') + f"\nPayment received via Razorpay: {payment_id}"
                    pfs.save()
        except Exception:
            pass

        return JsonResponse({'status': 'success'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Payment verification failed: {str(e)}'}, status=400)

# ============================================================================
# QUOTE & CONTACT FORMS
# ============================================================================

def quote_form(request, product_id):
    """Quote request form for a specific product"""
    product_item = get_object_or_404(ProductItem, id=product_id, is_active=True)
    
    if request.method == 'POST':
        # Handle quote form submission
        customer_name = request.POST.get('customer_name', '').strip()
        company_name = request.POST.get('company_name', '').strip()
        email = request.POST.get('email', '').strip()
        mobile = request.POST.get('mobile', '').strip()
        quantity = int(request.POST.get('quantity', 1))
        
        if not all([customer_name, email, mobile]):
            messages.error(request, 'Please fill out all required fields.')
            return redirect('quote_form', product_id=product_id)
        
        # Create or get customer
        customer, created = Customer.objects.get_or_create(
            email=email,
            defaults={
                'name': customer_name,
                'company_name': company_name,
                'mobile': mobile,
            }
        )
        
        # Calculate pricing
        basic_amount = product_item.basic_amount * quantity
        cgst = product_item.cgst * quantity
        sgst = product_item.sgst * quantity
        total_amount = basic_amount + cgst + sgst
        token_amount = product_item.token_amount or 0
        installing_charges = product_item.installing_charges or 0
        grand_total = total_amount + token_amount + installing_charges
        
        # Create quote submission
        quote = QuoteSubmission.objects.create(
            customer=customer,
            product_item=product_item,
            quantity=quantity,
            basic_amount=basic_amount,
            cgst=cgst,
            sgst=sgst,
            total_amount=total_amount,
            token_amount=token_amount,
            installing_charges=installing_charges,
            grand_total=grand_total,
            status='pending'
        )
        
        # Send admin notification email using common template
        admin_email_content = render_to_string('products/generic_form_email.html', {
            'form_title': 'Quote Request Submission',
            'form_subtitle': 'A new quote request has been submitted from fusiontec.com',
            'customer_info': {
                'name': customer_name,
                'email': email,
                'mobile': mobile,
                'company': company_name or 'Not provided'
            },
            'product_info': {
                'product': product_item.item_name,
                'quantity': quantity
            },
            'pricing_info': {
                'basic_amount': f'₹{basic_amount:,.2f}',
                'cgst': f'₹{cgst:,.2f}',
                'sgst': f'₹{sgst:,.2f}',
                'total_amount': f'₹{total_amount:,.2f}',
                'token_amount': f'₹{token_amount:,.2f}',
                'installing_charges': f'₹{installing_charges:,.2f}',
                'grand_total': f'₹{grand_total:,.2f}'
            },
            'form_name': 'Quote Request Form',
            'submission_id': quote.id,
            'customer_email': email
        })

        # Send customer thank you email using common template
        customer_thank_you_content = render_to_string('products/generic_form_thanks.html', {
            'thank_you_title': 'Thank You for Your Quote Request!',
            'thank_you_subtitle': 'We\'ve received your quote request and will prepare it shortly',
            'customer_name': customer_name,
            'form_type': 'quote request',
            'response_time': '24-48 hours',
            'submission_details': {
                'product': product_item.item_name,
                'quantity': quantity,
                'quote_id': quote.id
            },
            'next_steps': [
                'Our team will analyze your requirements',
                'We\'ll prepare a detailed quote with pricing',
                'You\'ll receive a comprehensive proposal',
                'We\'ll discuss implementation options and timeline'
            ],
            'important_info': 'Our team will review your requirements and get back to you with a detailed quote within 24-48 hours.',
            'support_email': 'support@fusiontec.com'
        })

        try:
            # Send to admin
            admin_email = EmailMessage(
                subject=f"[Fusiontec Quote Request] - {product_item.item_name} - {customer_name}",
                body=admin_email_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[settings.CONTACT_FORM_RECIPIENT],
            )
            admin_email.content_subtype = "html"
            
            # Embed logo inline using CID
            try:
                from django.contrib.staticfiles import finders
                logo_path = finders.find('products/img/fusiontec.png')
                if logo_path:
                    with open(logo_path, 'rb') as f:
                        admin_email.attach_inline('fusiontec_logo', f.read(), 'image/png')
            except Exception as _e:
                print(f"Inline logo attach failed (admin): {_e}")
            
            admin_email.send()

            # Send thank you to customer
            customer_email = EmailMessage(
                subject="Quote Request Received - Fusiontec",
                body=customer_thank_you_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            customer_email.content_subtype = "html"
            
            # Embed logo inline using CID
            try:
                from django.contrib.staticfiles import finders
                logo_path = finders.find('products/img/fusiontec.png')
                if logo_path:
                    with open(logo_path, 'rb') as f:
                        customer_email.attach_inline('fusiontec_logo', f.read(), 'image/png')
            except Exception as _e:
                print(f"Inline logo attach failed (customer): {_e}")
            
            customer_email.send()

            messages.success(request, 'Quote request submitted successfully! You will receive a confirmation email shortly.')
        except Exception as e:
            messages.error(request, f'Quote submitted but there was an error sending confirmation emails: {str(e)}')
        
        return redirect('quote_detail', quote_id=quote.id)
    
    context = {
        'product_item': product_item,
    }
    return render(request, 'products/quote_form.html', context)

def quote_detail(request, quote_id):
    """Show quote details"""
    quote = get_object_or_404(QuoteSubmission, id=quote_id)
    
    context = {
        'quote': quote,
    }
    return render(request, 'products/quote_detail.html', context)

def footer_contact_form(request):
    """Handle footer contact form submission"""
    print(f"Footer contact form called with method: {request.method}")
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()

        print(f"Form data received - Name: {name}, Email: {email}, Phone: {phone}, Subject: {subject}")

        if not all([name, email, phone, subject, message]):
            print("Missing required fields")
            messages.error(request, 'Please fill out all required fields.')
            return redirect(request.META.get('HTTP_REFERER', 'index'))

        # Save to database
        ContactSubmission.objects.create(
            name=name,
            email=email,
            phone=phone,
            subject=subject,
            message=message,
        )

        # Admin notification email
        email_html_content = render_to_string('products/contact_form_email.html', {
            'name': name,
            'email': email,
            'phone': phone,
            'subject': subject,
            'message': message,
        })

        # User thank-you email
        user_thank_you_content = render_to_string('products/contact_form_thanks.html', {
            'name': name,
        })

        try:
            # Send to admin (no logo)
            admin_email_msg = EmailMessage(
                subject=f"[Fusiontec Contact] - {subject}",
                body=email_html_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[settings.CONTACT_FORM_RECIPIENT],
            )
            admin_email_msg.content_subtype = "html"
            
            # Send admin email
            admin_email_msg.send()
            print(f"Admin email sent successfully to {settings.CONTACT_FORM_RECIPIENT}")

            # Send thank you to user (no logo)
            user_email_msg = EmailMessage(
                subject="Thanks for contacting Fusiontec!",
                body=user_thank_you_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            user_email_msg.content_subtype = "html"
            
            # Send user email
            user_email_msg.send()
            print(f"User email sent successfully to {email}")

            messages.success(request, 'Your message has been sent successfully! You will receive a confirmation email shortly.')
        except Exception as e:
            print(f"Email sending error: {str(e)}")
            messages.error(request, f'Message saved but there was an error sending confirmation emails: {str(e)}')

        # Redirect back to the referring page
        return redirect(request.META.get('HTTP_REFERER', 'index'))
    
    # If it's a GET request, redirect to home page
    return redirect('index')

def contact_form(request):
    """Contact form page"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()

        if not all([name, email, phone, subject, message]):
            messages.error(request, 'Please fill out all required fields.')
            return redirect('contact_form')

        # Save to database
        ContactSubmission.objects.create(
            name=name,
            email=email,
            phone=phone,
            subject=subject,
            message=message,
        )

        # Admin notification email
        email_html_content = render_to_string('products/contact_form_email.html', {
            'name': name,
            'email': email,
            'phone': phone,
            'subject': subject,
            'message': message,
        })

        # User thank-you email
        user_thank_you_content = render_to_string('products/contact_form_thanks.html', {
            'name': name,
        })

        try:
            # Send to admin
            admin_email_msg = EmailMessage(
                subject=f"[Fusiontec Contact] - {subject}",
                body=email_html_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[settings.CONTACT_FORM_RECIPIENT],
            )
            admin_email_msg.content_subtype = "html"
            
            # Embed logo inline using CID
            try:
                from django.contrib.staticfiles import finders
                logo_path = finders.find('products/img/fusiontec.png')
                if logo_path:
                    with open(logo_path, 'rb') as f:
                        admin_email_msg.attach_inline('fusiontec_logo', f.read(), 'image/png')
            except Exception as _e:
                print(f"Inline logo attach failed (admin): {_e}")
            
            # Send admin email
            admin_email_msg.send()
            print(f"Admin email sent successfully to {settings.CONTACT_FORM_RECIPIENT}")

            # Send thank you to user
            user_email_msg = EmailMessage(
                subject="Thanks for contacting Fusiontec!",
                body=user_thank_you_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            user_email_msg.content_subtype = "html"
            
            # Embed logo inline using CID
            try:
                from django.contrib.staticfiles import finders
                logo_path = finders.find('products/img/fusiontec.png')
                if logo_path:
                    with open(logo_path, 'rb') as f:
                        user_email_msg.attach_inline('fusiontec_logo', f.read(), 'image/png')
            except Exception as _e:
                print(f"Inline logo attach failed (user): {_e}")
            
            # Send user email
            user_email_msg.send()
            print(f"User email sent successfully to {email}")

            messages.success(request, 'Your message has been sent successfully! You will receive a confirmation email shortly.')
        except Exception as e:
            print(f"Email sending error: {str(e)}")
            messages.error(request, f'Message saved but there was an error sending confirmation emails: {str(e)}')

        # Redirect back to the referring page or home page
        referer = request.META.get('HTTP_REFERER')
        if referer and 'contact' not in referer:
            return redirect(referer)
        else:
            return redirect('index')
    
    # If it's a GET request, redirect to home page
    return redirect('index')

def submit_quote(request):
    """Handle quote form submission and send email using stored credentials"""
    if request.method == 'POST':
        try:
            # Get form data
            customer_name = request.POST.get('customer_name', '').strip()
            mobile = request.POST.get('mobile', '').strip()
            email = request.POST.get('email', '').strip()
            gst_number = request.POST.get('gst_number', '').strip()
            message = request.POST.get('message', '').strip()
            product_type_id = request.POST.get('product_type_id')
            
            # Debug logging
            print(f"Quote submission - product_type_id: {product_type_id}, customer: {customer_name}, email: {email}")
            
            # Validate required fields
            if not all([customer_name, mobile, email, gst_number, message]):
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'status': 'error', 'message': 'Please fill all required fields.'})
                messages.error(request, 'Please fill all required fields.')
                referer = request.META.get('HTTP_REFERER')
                return redirect(referer or 'index')
            
            # Get product type details - handle DSC case specially
            product_type = None
            is_dsc = False
            
            if product_type_id and product_type_id != '0':
                # Check if this is the special DSC identifier
                if product_type_id == 'dsc':
                    is_dsc = True
                    print("DSC submission detected via special identifier")
                    # For DSC submissions, create or get a DSC product type
                    product_type, created = ProductTypeMaster.objects.get_or_create(
                        prdt_desc='Digital Signature Certificate (DSC)',
                        defaults={
                            'prdt_desc': 'Digital Signature Certificate (DSC)',
                            'sender_email': 'dsc@fusiontec.com',
                            'app_password': 'pcjn sxte zvci tljs'
                        }
                    )
                    if created:
                        print(f"Created new DSC product type with ID: {product_type.id}")
                    else:
                        print(f"Found existing DSC product type with ID: {product_type.id}")
                else:
                    try:
                        product_type = ProductTypeMaster.objects.get(id=product_type_id)
                        print(f"Found product type by ID: {product_type.id} - {product_type.prdt_desc}")
                    except ProductTypeMaster.DoesNotExist:
                        print(f"Product type with ID {product_type_id} not found")
                        pass
            
            # If no valid product type found, check if this is a DSC submission
            if not product_type:
                # Check if this is coming from DSC section (either by referrer or by context)
                referer = request.META.get('HTTP_REFERER', '')
                is_dsc = (
                    '/dsc/' in referer or 
                    'dsc' in referer.lower() or
                    'digital' in referer.lower()
                )
                
                if is_dsc:
                    print("DSC submission detected via referrer")
                    # For DSC submissions, create or get a DSC product type
                    product_type, created = ProductTypeMaster.objects.get_or_create(
                        prdt_desc='Digital Signature Certificate (DSC)',
                        defaults={
                            'prdt_desc': 'Digital Signature Certificate (DSC)',
                            'sender_email': 'dsc@fusiontec.com',
                            'app_password': 'pcjn sxte zvci tljs'
                        }
                    )
                    if created:
                        print(f"Created new DSC product type with ID: {product_type.id}")
                    else:
                        print(f"Found existing DSC product type with ID: {product_type.id}")
                else:
                    # Try to find any product type as fallback
                    product_type = ProductTypeMaster.objects.first()
                    if not product_type:
                        print("No product types found in system")
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'status': 'error', 'message': 'No product types configured in system.'})
                        messages.error(request, 'No product types configured in system.')
                        referer = request.META.get('HTTP_REFERER')
                        return redirect(referer or 'index')
                    else:
                        print(f"Using fallback product type: {product_type.id} - {product_type.prdt_desc}")
            
            # Determine if this is a DSC submission
            if not is_dsc:
                try:
                    _name = (product_type.prdt_desc or '').strip().lower()
                    is_dsc = (
                        ('dsc' in _name) or
                        ('digital' in _name and 'signature' in _name) or
                        ('mudhra' in _name)
                    )
                    if is_dsc:
                        print(f"DSC submission detected via product name: {_name}")
                except Exception:
                    is_dsc = False
            
            print(f"Final determination - is_dsc: {is_dsc}, product_type: {product_type.id} - {product_type.prdt_desc}")
            
            # Create submission record
            from django.utils import timezone
            quote_request = QuoteRequest.objects.create(
                customer_name=customer_name,
                company_name='',  # Not collected in new form
                mobile=mobile,
                email=email,
                product_type=product_type,
                quantity=1,  # Default to 1
                address='',  # Not collected in new form
                state='',  # Not collected in new form
                district='',  # Not collected in new form
                pincode='',  # Not collected in new form
                gst_number=gst_number,
                additional_requirements=message,
                status='new'
            )
            
            print(f"Created quote request with ID: {quote_request.id}")
            
            # Set email credentials based on product type
            if is_dsc:
                sender_email_cfg = 'dsc@fusiontec.com'
                app_password_cfg = 'pcjn sxte zvci tljs'
                print("Using DSC email credentials")
            else:
                sender_email_cfg = (product_type.sender_email or '').strip()
                app_password_cfg = (product_type.app_password or '').strip()
                print(f"Using product type email credentials: {sender_email_cfg}")
            
            # Gmail app passwords are often spaced for readability; strip any spaces
            sanitized_password = app_password_cfg.replace(' ', '') if app_password_cfg else ''

            # Send email if email credentials are configured
            print(f"Checking email credentials: sender_email='{sender_email_cfg}', app_password='{'Yes' if sanitized_password else 'No'}'")
            if sender_email_cfg and sanitized_password:
                print(f"Attempting to send email from {sender_email_cfg} to {email}")
                try:
                    # Configure email backend with stored credentials
                    email_backend = EmailBackend(
                        host='smtp.gmail.com',
                        port=587,
                        username=sender_email_cfg,
                        password=sanitized_password,
                        use_tls=True,
                        fail_silently=False
                    )

                    # Prepare email content
                    subject = f"Enquiry Confirmation - {product_type.prdt_desc}"

                    # Render HTML email template
                    
                    email_context = {
                        'customer_name': customer_name,
                        'product_type': product_type.prdt_desc,
                        'mobile': mobile,
                        'email': email,
                        'gst_number': gst_number,
                        'message': message,
                        'enquiry_id': quote_request.id,
                    }
                    
                    email_body = render_to_string('products/email_templates/enquiry_confirmation.html', email_context)

                    # Send email to customer using custom backend
                    email_message = EmailMessage(
                        subject=subject,
                        body=email_body,
                        from_email=sender_email_cfg,
                        to=[email],
                        reply_to=[sender_email_cfg]
                    )
                    email_message.content_subtype = "html"  # Set content type to HTML
                    email_message.connection = email_backend
                    email_message.send()

                    # Send notification to admin
                    admin_subject = f"New Enquiry - {product_type.prdt_desc}"
                    
                    # Render admin HTML email template
                    admin_context = {
                        'customer_name': customer_name,
                        'product_type': product_type.prdt_desc,
                        'mobile': mobile,
                        'email': email,
                        'gst_number': gst_number,
                        'message': message,
                        'enquiry_id': quote_request.id,
                        'submission_time': timezone.now().strftime('%B %d, %Y at %I:%M %p'),
                        'admin_dashboard_url': f"{request.build_absolute_uri('/')}admin-dashboard/",
                        'settings_url': f"{request.build_absolute_uri('/')}admin-dashboard/",
                    }
                    
                    admin_body = render_to_string('products/email_templates/admin_notification.html', admin_context)

                    admin_email = EmailMessage(
                        subject=admin_subject,
                        body=admin_body,
                        from_email=sender_email_cfg,
                        to=['dsc@fusiontec.com'] if is_dsc else [settings.CONTACT_FORM_RECIPIENT]
                    )
                    admin_email.content_subtype = "html"  # Set content type to HTML
                    admin_email.connection = email_backend
                    admin_email.send()

                    # Update email status
                    quote_request.email_sent = True
                    quote_request.email_sent_at = timezone.now()
                    quote_request.save()

                    print("Email sent successfully")
                    messages.success(request, 'Enquiry submitted successfully! We will contact you soon.')

                except Exception as email_error:
                    # Log the email error but don't fail the submission
                    print(f"Email sending failed: {email_error}")
                    print(f"Email credentials: {sender_email_cfg}")
                    messages.success(request, 'Enquiry submitted successfully! We will contact you soon.')
            else:
                # No email credentials configured, just show success message
                print(f"No email credentials configured for product type: {product_type.prdt_desc}")
                messages.success(request, 'Enquiry submitted successfully! We will contact you soon.')

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'success'})
            referer = request.META.get('HTTP_REFERER')
            return redirect(referer or 'index')

        except Exception as e:
            print(f"Error in submit_quote: {str(e)}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'status': 'error', 'message': f'Error submitting quote: {str(e)}'})
            messages.error(request, f'Error submitting quote: {str(e)}')
            referer = request.META.get('HTTP_REFERER')
            return redirect(referer or 'index')

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'error', 'message': 'Invalid request'})
    referer = request.META.get('HTTP_REFERER')
    return redirect(referer or 'index')

# ============================================================================
# ADMIN FUNCTIONS
# ============================================================================

def custom_admin_logout(request):
    """Admin logout function"""
    if 'admin_logged_in' in request.session:
        del request.session['admin_logged_in']
    if 'admin_username' in request.session:
        del request.session['admin_username']
    messages.success(request, "Logout successfully")
    return redirect('/')

def custom_admin_login(request):
    """Admin login function"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Simple admin check (you should implement proper authentication)
        if username == 'admin' and password == 'admin':
            request.session['admin_logged_in'] = True
            request.session['admin_username'] = username
            messages.success(request, 'Login successful!')
            return redirect('custom_admin_dashboard')
        else:
            messages.error(request, 'Invalid credentials!')
    
    return render(request, 'products/admin/login.html')

@admin_required
def custom_admin_dashboard(request):
    """Admin dashboard with comprehensive data"""
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Get current date and calculate date ranges
    now = timezone.now()
    today = now.date()
    this_month = now.replace(day=1)
    last_month = (this_month - timedelta(days=1)).replace(day=1)
    
    # Basic counts
    total_customers = Customer.objects.count()
    total_quotes = QuoteSubmission.objects.count()
    total_contacts = ContactSubmission.objects.count()
    total_products = ProductMasterV2.objects.count()
    
    # Monthly statistics
    this_month_quotes = QuoteSubmission.objects.filter(
        created_at__gte=this_month
    ).count()
    this_month_contacts = ContactSubmission.objects.filter(
        created_at__gte=this_month
    ).count()
    this_month_customers = Customer.objects.filter(
        created_at__gte=this_month
    ).count()
    
    last_month_quotes = QuoteSubmission.objects.filter(
        created_at__gte=last_month,
        created_at__lt=this_month
    ).count()
    last_month_contacts = ContactSubmission.objects.filter(
        created_at__gte=last_month,
        created_at__lt=this_month
    ).count()
    last_month_customers = Customer.objects.filter(
        created_at__gte=last_month,
        created_at__lt=this_month
    ).count()
    
    # Calculate growth percentages
    quote_growth = ((this_month_quotes - last_month_quotes) / max(last_month_quotes, 1)) * 100 if last_month_quotes > 0 else 0
    contact_growth = ((this_month_contacts - last_month_contacts) / max(last_month_contacts, 1)) * 100 if last_month_contacts > 0 else 0
    customer_growth = ((this_month_customers - last_month_customers) / max(last_month_customers, 1)) * 100 if last_month_customers > 0 else 0
    
    # Quote status distribution
    quote_status_counts = QuoteSubmission.objects.values('status').annotate(
        count=Count('id')
    ).order_by('status')
    
    # Product performance (top products by quote count)
    top_products = QuoteSubmission.objects.values(
        'product_item__item_name'
    ).annotate(
        quote_count=Count('id'),
        total_amount=Sum('grand_total')
    ).order_by('-quote_count')[:5]
    
    # Recent activities (last 7 days)
    week_ago = now - timedelta(days=7)
    recent_quotes = QuoteSubmission.objects.select_related(
        'customer', 'product_item'
    ).filter(
        created_at__gte=week_ago
    ).order_by('-created_at')[:10]
    
    recent_contacts = ContactSubmission.objects.filter(
        created_at__gte=week_ago
    ).order_by('-created_at')[:10]
    
    # Today's activities
    today_quotes = QuoteSubmission.objects.filter(
        created_at__date=today
    ).count()
    today_contacts = ContactSubmission.objects.filter(
        created_at__date=today
    ).count()
    today_customers = Customer.objects.filter(
        created_at__date=today
    ).count()
    
    # Revenue statistics
    total_revenue = QuoteSubmission.objects.filter(
        status='approved'
    ).aggregate(
        total=Sum('grand_total')
    )['total'] or 0
    
    this_month_revenue = QuoteSubmission.objects.filter(
        status='approved',
        created_at__gte=this_month
    ).aggregate(
        total=Sum('grand_total')
    )['total'] or 0
    
    last_month_revenue = QuoteSubmission.objects.filter(
        status='approved',
        created_at__gte=last_month,
        created_at__lt=this_month
    ).aggregate(
        total=Sum('grand_total')
    )['total'] or 0
    
    revenue_growth = ((this_month_revenue - last_month_revenue) / max(last_month_revenue, 1)) * 100 if last_month_revenue > 0 else 0
    
    # Customer location distribution (top states)
    top_states = Customer.objects.exclude(
        state__isnull=True
    ).exclude(
        state=''
    ).values('state').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Weekly trend data for charts
    weekly_data = []
    for i in range(7):
        date = today - timedelta(days=i)
        day_quotes = QuoteSubmission.objects.filter(
            created_at__date=date
        ).count()
        day_contacts = ContactSubmission.objects.filter(
            created_at__date=date
        ).count()
        weekly_data.append({
            'date': date.strftime('%b %d'),
            'quotes': day_quotes,
            'contacts': day_contacts
        })
    weekly_data.reverse()
    
    context = {
        # Basic counts
        'total_customers': total_customers,
        'total_quotes': total_quotes,
        'total_contacts': total_contacts,
        'total_products': total_products,
        
        # Monthly statistics
        'this_month_quotes': this_month_quotes,
        'this_month_contacts': this_month_contacts,
        'this_month_customers': this_month_customers,
        
        # Growth percentages
        'quote_growth': round(quote_growth, 1),
        'contact_growth': round(contact_growth, 1),
        'customer_growth': round(customer_growth, 1),
        
        # Quote status distribution
        'quote_status_counts': quote_status_counts,
        
        # Top products
        'top_products': top_products,
        
        # Recent activities
        'recent_quotes': recent_quotes,
        'recent_contacts': recent_contacts,
        
        # Today's activities
        'today_quotes': today_quotes,
        'today_contacts': today_contacts,
        'today_customers': today_customers,
        
        # Revenue statistics
        'total_revenue': total_revenue,
        'this_month_revenue': this_month_revenue,
        'revenue_growth': round(revenue_growth, 1),
        
        # Location data
        'top_states': top_states,
        
        # Chart data
        'weekly_data': weekly_data,
        
        # Chart data for enhanced dashboard
        'monthly_labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
        'monthly_quotes': [this_month_quotes, last_month_quotes, 0, 0, 0, 0],
        'monthly_contacts': [this_month_contacts, last_month_contacts, 0, 0, 0, 0],
        'product_labels': ['Tally', 'e-Mudhra', 'FusionTec', 'Business'],
        'product_data': [25, 30, 20, 25],
        'revenue_labels': ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'],
        'revenue_data': [this_month_revenue, last_month_revenue, 0, 0, 0, 0],
        
        # Date info
        'today': today.strftime('%B %d, %Y'),
        'this_month': this_month.strftime('%B %Y'),
    }
    return render(request, 'products/admin/dashboard.html', context)


@admin_required
def dashboard_data_api(request):
    """API endpoint for dashboard data updates"""
    from django.http import JsonResponse
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Get current date and calculate date ranges
    now = timezone.now()
    today = now.date()
    this_month = now.replace(day=1)
    last_month = (this_month - timedelta(days=1)).replace(day=1)
    
    # Basic counts
    total_customers = Customer.objects.count()
    total_quotes = QuoteSubmission.objects.count()
    total_contacts = ContactSubmission.objects.count()
    total_products = ProductMasterV2.objects.count()
    
    # Monthly statistics
    this_month_quotes = QuoteSubmission.objects.filter(
        created_at__gte=this_month
    ).count()
    this_month_contacts = ContactSubmission.objects.filter(
        created_at__gte=this_month
    ).count()
    
    # Recent activities
    recent_quotes = QuoteSubmission.objects.select_related('customer').order_by('-created_at')[:5]
    recent_contacts = ContactSubmission.objects.order_by('-created_at')[:5]
    
    # Format recent activities
    recent_quotes_data = []
    for quote in recent_quotes:
        recent_quotes_data.append({
            'title': f"Quote from {quote.customer.name}",
            'time': quote.created_at.strftime('%b %d, %H:%M'),
            'icon': 'file-invoice'
        })
    
    recent_contacts_data = []
    for contact in recent_contacts:
        recent_contacts_data.append({
            'title': f"Contact from {contact.name}",
            'time': contact.created_at.strftime('%b %d, %H:%M'),
            'icon': 'envelope'
        })
    
    data = {
        'total_customers': total_customers,
        'total_quotes': total_quotes,
        'total_contacts': total_contacts,
        'total_products': total_products,
        'this_month_quotes': this_month_quotes,
        'this_month_contacts': this_month_contacts,
        'recent_quotes': recent_quotes_data,
        'recent_contacts': recent_contacts_data,
    }
    
    return JsonResponse(data)

@admin_required
def custom_admin_contacts(request):
    """Admin contacts management"""
    contacts = ContactSubmission.objects.all().order_by('-created_at')
    
    # Pagination
    paginator = Paginator(contacts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    return render(request, 'products/admin/contacts.html', context)

@admin_required
def custom_admin_quotes(request):
    """Admin quotes management"""
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_status':
            quote_id = request.POST.get('quote_id')
            new_status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            try:
                quote = QuoteSubmission.objects.get(id=quote_id)
                quote.status = new_status
                quote.notes = notes
                quote.save()
                messages.success(request, f'Quote #{quote_id} status updated to {new_status}.')
            except QuoteSubmission.DoesNotExist:
                messages.error(request, 'Quote not found.')
            except Exception as e:
                messages.error(request, f'Failed to update quote: {e}')
            
            return redirect('custom_admin_quotes')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status_filter', '')
    
    # Build queryset with select_related for performance
    quotes = QuoteSubmission.objects.select_related(
        'customer', 
        'product_item__product_type__product_master'
    ).all().order_by('-created_at')
    
    # Apply search filter
    if search_query:
        quotes = quotes.filter(
            Q(customer__name__icontains=search_query) |
            Q(customer__mobile__icontains=search_query) |
            Q(customer__email__icontains=search_query) |
            Q(product_item__item_name__icontains=search_query) |
            Q(product_item__product_type__type_name__icontains=search_query)
        )
    
    # Apply status filter
    if status_filter:
        quotes = quotes.filter(status=status_filter)
    
    # Pagination
    paginator = Paginator(quotes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'products/admin/quotes.html', context)

@admin_required
def custom_admin_products(request):
    """Admin management for the NEW simple product tables.

    - Create ProductTypeMaster (if form_type = 'type')
    - Create ProductMasterV2 (if form_type = 'product')
    - Create ProductSubMaster (if form_type = 'subproduct')
    - Edit/Delete functionality
    """
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        action = request.POST.get('action')
        
        try:
            if action == 'delete_type':
                type_id = request.POST.get('type_id')
                ProductTypeMaster.objects.get(id=type_id).delete()
                messages.success(request, 'Product Type deleted.')
                return redirect('custom_admin_products')
            elif action == 'delete_product':
                product_id = request.POST.get('product_id')
                ProductMasterV2.objects.get(id=product_id).delete()
                messages.success(request, 'Product deleted.')
                return redirect('custom_admin_products')
            elif form_type == 'edit_type':
                type_id = request.POST.get('type_id')
                prdt_desc = (request.POST.get('prdt_desc') or '').strip()
                image = request.FILES.get('image')
                sender_email = (request.POST.get('sender_email') or '').strip()
                app_password = (request.POST.get('app_password') or '').strip()
                
                if not type_id or not prdt_desc:
                    messages.error(request, 'Type ID and description are required.')
                else:
                    try:
                        product_type = ProductTypeMaster.objects.get(id=int(type_id))
                        product_type.prdt_desc = prdt_desc
                        product_type.sender_email = sender_email if sender_email else None
                        product_type.app_password = app_password if app_password else None
                        if image:
                            product_type.image = image
                        product_type.save()
                        messages.success(request, 'Product Type updated.')
                    except ProductTypeMaster.DoesNotExist:
                        messages.error(request, 'Product Type not found.')
                    except Exception as exc:
                        messages.error(request, f'Failed to update: {exc}')
                return redirect('custom_admin_products')
            elif form_type == 'type':
                prdt_desc = (request.POST.get('prdt_desc') or '').strip()
                image = request.FILES.get('image')
                sender_email = (request.POST.get('sender_email') or '').strip()
                app_password = (request.POST.get('app_password') or '').strip()
                if not prdt_desc:
                    messages.error(request, 'Type description is required.')
                else:
                    ProductTypeMaster.objects.create(
                        prdt_desc=prdt_desc, 
                        image=image,
                        sender_email=sender_email if sender_email else None,
                        app_password=app_password if app_password else None
                    )
                    messages.success(request, 'Product Type created.')
                    return redirect('custom_admin_products')
            elif form_type == 'product':
                type_id = request.POST.get('product_type')
                prdt_desc = (request.POST.get('prdt_desc') or '').strip()
                if not type_id or not prdt_desc:
                    messages.error(request, 'Type and Product description are required.')
                else:
                    pt = ProductTypeMaster.objects.get(id=int(type_id))
                    ProductMasterV2.objects.create(product_type=pt, prdt_desc=prdt_desc)
                    messages.success(request, 'Product created.')
                    return redirect('custom_admin_products')
            elif form_type == 'subproduct':
                product_id = request.POST.get('product_id')
                subprdt_desc = (request.POST.get('subprdt_desc') or '').strip()
                if not product_id or not subprdt_desc:
                    messages.error(request, 'Product and Sub Product description are required.')
                else:
                    prod = ProductMasterV2.objects.get(id=int(product_id))
                    ProductSubMaster.objects.create(product=prod, subprdt_desc=subprdt_desc)
                    messages.success(request, 'Sub Product created.')
                    return redirect('custom_admin_products')
        except Exception as exc:
            messages.error(request, f'Failed to save: {exc}')

    types = ProductTypeMaster.objects.all().order_by('id')
    
    # Handle filtering by product type
    selected_type = request.GET.get('type_filter', '')
    products = ProductMasterV2.objects.select_related('product_type').all().order_by('id')
    
    print(f"DEBUG: selected_type = '{selected_type}', type = {type(selected_type)}")
    
    if selected_type:
        try:
            products = products.filter(product_type_id=int(selected_type))
            print(f"DEBUG: Filtered products count: {products.count()}")
        except (ValueError, TypeError) as e:
            print(f"DEBUG: Filter error: {e}")
            # If conversion fails, don't filter
            pass

    context = {
        'simple_types': types,
        'simple_products': products,
        'selected_type': selected_type,
        'sub_products': ProductSubMaster.objects.select_related('product').all().order_by('id'),
    }
    return render(request, 'products/admin/products.html', context)

@admin_required
def custom_admin_new_products(request):
    """Admin management for New Products (promotional images)"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        try:
            if action == 'delete':
                product_id = request.POST.get('product_id')
                NewProduct.objects.get(id=product_id).delete()
                messages.success(request, 'New Product deleted successfully.')
                return redirect('custom_admin_new_products')
            elif action == 'edit':
                product_id = request.POST.get('product_id')
                title = request.POST.get('title', '').strip()
                description = request.POST.get('description', '').strip()
                is_active = request.POST.get('is_active') == 'on'
                display_order = int(request.POST.get('display_order', 0))
                image = request.FILES.get('image')
                
                product = NewProduct.objects.get(id=product_id)
                product.title = title if title else None
                product.description = description if description else None
                product.is_active = is_active
                product.display_order = display_order
                if image:
                    product.image = image
                product.save()
                messages.success(request, 'New Product updated successfully.')
                return redirect('custom_admin_new_products')
            elif action == 'create':
                title = request.POST.get('title', '').strip()
                description = request.POST.get('description', '').strip()
                is_active = request.POST.get('is_active') == 'on'
                display_order = int(request.POST.get('display_order', 0))
                image = request.FILES.get('image')
                
                if not image:
                    messages.error(request, 'Image is required.')
                else:
                    NewProduct.objects.create(
                        title=title if title else None,
                        description=description if description else None,
                        image=image,
                        is_active=is_active,
                        display_order=display_order
                    )
                    messages.success(request, 'New Product created successfully.')
                    return redirect('custom_admin_new_products')
        except Exception as exc:
            messages.error(request, f'Error: {exc}')
    
    new_products = NewProduct.objects.all().order_by('display_order', '-created_at')
    context = {
        'new_products': new_products,
    }
    return render(request, 'products/admin/new_products.html', context)

@admin_required
def admin_product_types(request, master_id):
    """Deprecated in new schema: keep route for compatibility (no-op redirect)."""
    return redirect('custom_admin_products')

    if request.method == 'POST':
        type_code = (request.POST.get('type_code') or '').strip()
        type_name = (request.POST.get('type_name') or '').strip()
        description = (request.POST.get('description') or '').strip()
        is_active = request.POST.get('is_active') == 'on'
        display_order = request.POST.get('display_order') or '0'

        if not type_code or not type_name:
            messages.error(request, 'Type Code and Type Name are required.')
        else:
            try:
                ProductType.objects.create(
                    product_master=master,
                    type_code=type_code,
                    type_name=type_name,
                    description=description or None,
                    is_active=is_active,
                    display_order=int(display_order),
                )
                messages.success(request, 'Product type created successfully.')
                return redirect('admin_product_types', master_id=master.id)
            except Exception as exc:
                messages.error(request, f'Failed to create product type: {exc}')

    types_qs = master.product_types.all().order_by('display_order', 'type_name')
    context = {
        'master': master,
        'types': types_qs,
    }
    return render(request, 'products/admin/product_types.html', context)

@admin_required
def admin_product_items(request, type_id):
    """Deprecated in new schema: keep route for compatibility (no-op redirect)."""
    return redirect('custom_admin_products')

    if request.method == 'POST':
        item_code = (request.POST.get('item_code') or '').strip()
        item_name = (request.POST.get('item_name') or '').strip()
        item_category = (request.POST.get('item_category') or 'product').strip()
        description = (request.POST.get('description') or '').strip()
        features = (request.POST.get('features') or '').strip()
        is_active = request.POST.get('is_active') == 'on'
        display_order = int((request.POST.get('display_order') or '0').strip())

        # pricing
        def to_decimal(val):
            try:
                return None if val is None or val == '' else float(val)
            except Exception:
                return None

        basic_amount = to_decimal(request.POST.get('basic_amount'))
        cgst = to_decimal(request.POST.get('cgst')) or 0
        sgst = to_decimal(request.POST.get('sgst')) or 0
        token_name = (request.POST.get('token_name') or '').strip() or None
        token_amount = to_decimal(request.POST.get('token_amount'))
        installing_charges = to_decimal(request.POST.get('installing_charges'))
        billing_cycle = (request.POST.get('billing_cycle') or '').strip() or None
        old_price = to_decimal(request.POST.get('old_price'))

        if not item_code or not item_name:
            messages.error(request, 'Item Code and Item Name are required.')
        else:
            try:
                ProductItem.objects.create(
                    product_type=ptype,
                    item_code=item_code,
                    item_name=item_name,
                    item_category=item_category,
                    description=description or None,
                    features=features or None,
                    is_active=is_active,
                    display_order=display_order,
                    basic_amount=basic_amount,
                    cgst=cgst,
                    sgst=sgst,
                    token_name=token_name,
                    token_amount=token_amount,
                    installing_charges=installing_charges,
                    billing_cycle=billing_cycle,
                    old_price=old_price,
                )
                messages.success(request, 'Product item created successfully.')
                return redirect('admin_product_items', type_id=ptype.id)
            except Exception as exc:
                messages.error(request, f'Failed to create item: {exc}')

    items = ptype.product_items.all().order_by('display_order', 'item_name')
    context = {
        'ptype': ptype,
        'items': items,
    }
    return render(request, 'products/admin/product_items.html', context)

@admin_required
def admin_rate_cards(request, item_id):
    """Manage rate cards for ProductSubMaster by ID."""
    sub_product = get_object_or_404(ProductSubMaster, id=item_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'delete_rate':
            rate_id = request.POST.get('rate_id')
            try:
                RateCardEntry.objects.get(id=rate_id).delete()
                messages.success(request, 'Rate card deleted.')
                return redirect('admin_rate_cards', item_id=sub_product.id)
            except Exception as exc:
                messages.error(request, f'Failed to delete rate card: {exc}')
        elif action == 'edit_rate':
            # Handle edit - smart editing based on date
            rate_id = request.POST.get('rate_id')
            rate_date = (request.POST.get('rate_date') or '').strip()
            base_amount = (request.POST.get('base_amount') or '').strip()
            gst_percent = (request.POST.get('gst_percent') or '').strip()
            
            # Token fields
            token_desc = (request.POST.get('token_desc') or '').strip()
            token_amount = (request.POST.get('token_amount') or '').strip()
            token_gst_percent = (request.POST.get('token_gst_percent') or '').strip()
            
            # Installation fields
            installation_charge = (request.POST.get('installation_charge') or '').strip()
            installation_gst_percent = (request.POST.get('installation_gst_percent') or '').strip()

            from datetime import datetime
            try:
                rate_dt = datetime.strptime(rate_date, '%Y-%m-%d').date()
            except Exception:
                messages.error(request, 'Invalid Rate Date. Use YYYY-MM-DD format.')
                return redirect('admin_rate_cards', item_id=sub_product.id)

            try:
                base_amt = float(base_amount)
                gst_pct = float(gst_percent or 0)
                token_amt = float(token_amount or 0)
                token_gst_pct = float(token_gst_percent or 18)
                install_charge = float(installation_charge or 0)
                install_gst_pct = float(installation_gst_percent or 18)
            except Exception:
                messages.error(request, 'Invalid amounts provided.')
                return redirect('admin_rate_cards', item_id=product.id)

            try:
                # Get the existing rate card entry
                existing_rate = RateCardEntry.objects.get(id=rate_id)
                
                # Check if the date is the same as the original
                if existing_rate.rate_date == rate_dt:
                    # Same date - update the existing entry
                    existing_rate.base_amt = base_amt
                    existing_rate.gst_percent = gst_pct
                    existing_rate.token_desc = token_desc
                    existing_rate.token_amount = token_amt
                    existing_rate.token_gst_percent = token_gst_pct
                    existing_rate.installation_charge = install_charge
                    existing_rate.installation_gst_percent = install_gst_pct
                    existing_rate.save()
                    messages.success(request, 'Rate card updated successfully (same date).')
                else:
                    # Different date - create a new entry for history
                    RateCardEntry.objects.create(
                        sub_product=sub_product,
                        rate_date=rate_dt,
                        base_amt=base_amt,
                        gst_percent=gst_pct,
                        token_desc=token_desc,
                        token_amount=token_amt,
                        token_gst_percent=token_gst_pct,
                        installation_charge=install_charge,
                        installation_gst_percent=install_gst_pct,
                    )
                    messages.success(request, 'Rate card updated successfully (new entry created for new date).')
                
                return redirect('admin_rate_cards', item_id=sub_product.id)
            except RateCardEntry.DoesNotExist:
                messages.error(request, 'Rate card not found.')
                return redirect('admin_rate_cards', item_id=sub_product.id)
            except Exception as exc:
                messages.error(request, f'Failed to update rate card: {exc}')
                return redirect('admin_rate_cards', item_id=sub_product.id)
        else:
            # Handle new rate card creation
            rate_date = (request.POST.get('rate_date') or '').strip()
            base_amount = (request.POST.get('base_amount') or '').strip()
            gst_percent = (request.POST.get('gst_percent') or '').strip()
            
            # Token fields
            token_desc = (request.POST.get('token_desc') or '').strip()
            token_amount = (request.POST.get('token_amount') or '').strip()
            token_gst_percent = (request.POST.get('token_gst_percent') or '').strip()
            
            # Installation fields
            installation_charge = (request.POST.get('installation_charge') or '').strip()
            installation_gst_percent = (request.POST.get('installation_gst_percent') or '').strip()

            from datetime import datetime
            try:
                rate_dt = datetime.strptime(rate_date, '%Y-%m-%d').date()
            except Exception:
                messages.error(request, 'Invalid Rate Date. Use YYYY-MM-DD format.')
                return redirect('admin_rate_cards', item_id=sub_product.id)

            try:
                base_amt = float(base_amount)
                gst_pct = float(gst_percent or 0)
                token_amt = float(token_amount or 0)
                token_gst_pct = float(token_gst_percent or 18)
                install_charge = float(installation_charge or 0)
                install_gst_pct = float(installation_gst_percent or 18)
            except Exception:
                messages.error(request, 'Invalid amounts provided.')
                return redirect('admin_rate_cards', item_id=product.id)

            try:
                RateCardEntry.objects.create(
                    sub_product=sub_product,
                    rate_date=rate_dt,
                    base_amt=base_amt,
                    gst_percent=gst_pct,
                    token_desc=token_desc,
                    token_amount=token_amt,
                    token_gst_percent=token_gst_pct,
                    installation_charge=install_charge,
                    installation_gst_percent=install_gst_pct,
                )
                messages.success(request, 'Rate card added successfully.')
                return redirect('admin_rate_cards', item_id=sub_product.id)
            except Exception as exc:
                messages.error(request, f'Failed to add rate card: {exc}')

    cards = sub_product.rate_cards.all().order_by('-rate_date')
    context = {
        'simple_product': sub_product,
        'cards': cards,
    }
    return render(request, 'products/admin/rate_cards.html', context)

# ============================================================================
# CUSTOM ADMIN VIEWS (Additional)
# ============================================================================

@admin_required
def custom_admin_tally_submissions(request):
    """Custom admin for Tally submissions - DEPRECATED"""
    context = {
        'page_obj': [],
        'search_query': '',
        'deprecated': True,
        'message': 'The old Tally submissions system has been deprecated. Please use the new unified quotes system.'
    }
    return render(request, 'products/admin/tally_submissions.html', context)

@admin_required
def custom_admin_emudhra_submissions(request):
    """Custom admin for e-Mudhra submissions - DEPRECATED"""
    context = {
        'page_obj': [],
        'search_query': '',
        'deprecated': True,
        'message': 'The old e-Mudhra submissions system has been deprecated. Please use the new unified quotes system.'
    }
    return render(request, 'products/admin/emudhra_submissions.html', context)

@admin_required
def custom_admin_fusiontec_submissions(request):
    """Custom admin for Fusiontec submissions - DEPRECATED"""
    context = {
        'page_obj': [],
        'search_query': '',
        'deprecated': True,
        'message': 'The old Fusiontec submissions system has been deprecated. Please use the new unified quotes system.'
    }
    return render(request, 'products/admin/fusiontec_submissions.html', context)

@admin_required
def custom_admin_biz_submissions(request):
    """Custom admin for Business Intelligence submissions"""
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'delete_submission':
            submission_id = request.POST.get('submission_id')
            try:
                quote = QuoteSubmission.objects.get(id=submission_id)
                quote.delete()
                messages.success(request, f'Quote #{submission_id} deleted successfully.')
            except QuoteSubmission.DoesNotExist:
                messages.error(request, 'Quote not found.')
            except Exception as e:
                messages.error(request, f'Failed to delete quote: {e}')
            
            return redirect('custom_admin_biz_submissions')
    
    # Get search query
    search_query = request.GET.get('search', '')
    
    # Fetch quotes related to Business Intelligence products
    quotes = QuoteSubmission.objects.select_related(
        'customer', 
        'product_item__product_type__product_master'
    ).filter(
        product_item__product_type__product_master__product_code='biz'
    ).order_by('-created_at')
    
    # Apply search filter
    if search_query:
        quotes = quotes.filter(
            Q(customer__name__icontains=search_query) |
            Q(customer__mobile__icontains=search_query) |
            Q(customer__email__icontains=search_query) |
            Q(product_item__item_name__icontains=search_query) |
            Q(product_item__product_type__type_name__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(quotes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'products/admin/biz_submissions.html', context)

@admin_required
def custom_admin_payments(request):
    """Custom admin for payment transactions"""
    search_query = request.GET.get('search', '')
    from .models import PaymentTransaction
    payments = PaymentTransaction.objects.all()
    
    if search_query:
        payments = payments.filter(
            Q(customer__name__icontains=search_query) |
            Q(razorpay_payment_id__icontains=search_query) |
            Q(razorpay_order_id__icontains=search_query) |
            Q(customer__email__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(payments.order_by('-created_at'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'products/admin/payments.html', context)

@admin_required
def custom_admin_applicants(request):
    """Custom admin for applicant submissions"""
    search_query = request.GET.get('search', '')
    applicants = Applicant.objects.all()
    
    if search_query:
        applicants = applicants.filter(
            Q(name__icontains=search_query) |
            Q(mobile_number__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(applicants.order_by('-created_at'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'products/admin/applicants.html', context)

@admin_required
def custom_admin_settings(request):
    """Custom admin for payment settings"""
    from .models import PaymentSettings
    
    # Get payment settings by type
    razorpay_settings = PaymentSettings.objects.filter(setting_type='razorpay', is_active=True).first()
    qr_settings = PaymentSettings.objects.filter(setting_type='qr_code', is_active=True).first()
    bank_settings = PaymentSettings.objects.filter(setting_type='bank_transfer', is_active=True).first()
    
    context = {
        'razorpay_info': [razorpay_settings] if razorpay_settings else [],
        'qr_info': [qr_settings] if qr_settings else [],
        'bank_info': [bank_settings] if bank_settings else [],
    }
    return render(request, 'products/admin/settings.html', context)

@admin_required
def admin_dsc_prices(request):
    """Admin page to view/update DSC dynamic prices."""
    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        row_id = request.POST.get('row_id')
        class_type = request.POST.get('class_type')
        user_type = request.POST.get('user_type')
        cert_type = request.POST.get('cert_type')
        validity = request.POST.get('validity')
        dsc_charge = request.POST.get('dsc_charge')
        token_amount = request.POST.get('token_amount') or '0'
        installation_charge = request.POST.get('installation_charge') or '0'
        gst_percent = request.POST.get('gst_percent') or '18'
        surcharge = request.POST.get('outside_india_surcharge') or '0'
        is_active = request.POST.get('is_active') == 'on'

        try:
            if action == 'delete' and row_id:
                DscPrice.objects.filter(id=row_id).delete()
                messages.success(request, 'Price row deleted.')
                return redirect('admin_dsc_prices')

            if row_id:
                row = DscPrice.objects.get(id=row_id)
                row.class_type = class_type
                row.user_type = user_type
                row.cert_type = cert_type
                row.validity = validity
                row.dsc_charge = float(dsc_charge or 0)
                row.token_amount = float(token_amount or 0)
                row.installation_charge = float(installation_charge or 0)
                row.gst_percent = float(gst_percent or 0)
                row.outside_india_surcharge = float(surcharge or 0)
                row.is_active = is_active
                row.save()
                messages.success(request, 'Price updated.')
            else:
                DscPrice.objects.create(
                    class_type=class_type,
                    user_type=user_type,
                    cert_type=cert_type,
                    validity=validity,
                    dsc_charge=float(dsc_charge or 0),
                    token_amount=float(token_amount or 0),
                    installation_charge=float(installation_charge or 0),
                    gst_percent=float(gst_percent or 0),
                    outside_india_surcharge=float(surcharge or 0),
                    is_active=is_active,
                )
                messages.success(request, 'Price added.')
            return redirect('admin_dsc_prices')
        except Exception as exc:
            messages.error(request, f'Failed to save price: {exc}')

    prices = DscPrice.objects.order_by('class_type','user_type','cert_type','validity')
    context = { 'prices': prices }
    return render(request, 'products/admin/dsc_prices.html', context)

@admin_required
def custom_admin_form_submissions(request):
    """Admin dashboard for product form submissions"""
    submissions = ProductFormSubmission.objects.all().order_by('-created_at')
    
    # Filtering
    status_filter = request.GET.get('status', '')
    product_filter = request.GET.get('product', '')
    date_filter = request.GET.get('date', '')
    
    if status_filter:
        submissions = submissions.filter(status=status_filter)
    
    if product_filter:
        submissions = submissions.filter(product_id__id=product_filter)
    
    if date_filter:
        submissions = submissions.filter(created_at__date=date_filter)
    
    # Pagination
    paginator = Paginator(submissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get unique products for filter
    products = ProductMasterV2.objects.all().order_by('prdt_desc')
    
    # Statistics
    total_submissions = ProductFormSubmission.objects.count()
    new_submissions = ProductFormSubmission.objects.filter(status='new').count()
    reviewed_submissions = ProductFormSubmission.objects.filter(status='reviewed').count()
    approved_submissions = ProductFormSubmission.objects.filter(status='approved').count()
    converted_submissions = ProductFormSubmission.objects.filter(status='converted').count()
    
    context = {
        'page_obj': page_obj,
        'total_submissions': total_submissions,
        'new_submissions': new_submissions,
        'reviewed_submissions': reviewed_submissions,
        'approved_submissions': approved_submissions,
        'converted_submissions': converted_submissions,
        'products': products,
        'status_filter': status_filter,
        'product_filter': product_filter,
        'date_filter': date_filter,
    }
    
    return render(request, 'products/admin/form_submissions.html', context)

@admin_required
def custom_admin_quote_requests(request):
    """Admin dashboard for quote requests"""
    from .models import QuoteRequest
    quote_requests = QuoteRequest.objects.all().order_by('-created_at')
    
    # Filtering
    status_filter = request.GET.get('status', '')
    product_type_filter = request.GET.get('product_type', '')
    date_filter = request.GET.get('date', '')
    
    if status_filter:
        quote_requests = quote_requests.filter(status=status_filter)
    
    if product_type_filter:
        quote_requests = quote_requests.filter(product_type_id=product_type_filter)
    
    if date_filter:
        quote_requests = quote_requests.filter(created_at__date=date_filter)
    
    # Pagination
    paginator = Paginator(quote_requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get unique product types for filter
    product_types = ProductTypeMaster.objects.all().order_by('prdt_desc')
    
    # Statistics
    total_requests = QuoteRequest.objects.count()
    new_requests = QuoteRequest.objects.filter(status='new').count()
    reviewed_requests = QuoteRequest.objects.filter(status='reviewed').count()
    quoted_requests = QuoteRequest.objects.filter(status='quoted').count()
    accepted_requests = QuoteRequest.objects.filter(status='accepted').count()
    
    context = {
        'page_obj': page_obj,
        'total_requests': total_requests,
        'new_requests': new_requests,
        'reviewed_requests': reviewed_requests,
        'quoted_requests': quoted_requests,
        'accepted_requests': accepted_requests,
        'product_types': product_types,
        'status_filter': status_filter,
        'product_type_filter': product_type_filter,
        'date_filter': date_filter,
    }
    
    return render(request, 'products/admin/quote_requests.html', context)

@admin_required
def quote_request_detail(request, request_id):
    """API endpoint to get quote request details"""
    from .models import QuoteRequest
    try:
        quote_request = QuoteRequest.objects.get(id=request_id)
        
        # Render the details as HTML
        html_content = f"""
        <div class="row">
            <div class="col-md-6">
                <h6 class="text-primary">Customer Information</h6>
                <table class="table table-sm">
                    <tr><td><strong>Name:</strong></td><td>{quote_request.customer_name}</td></tr>
                    <tr><td><strong>Company:</strong></td><td>{quote_request.company_name or 'Not specified'}</td></tr>
                    <tr><td><strong>Mobile:</strong></td><td>{quote_request.mobile}</td></tr>
                    <tr><td><strong>Email:</strong></td><td>{quote_request.email}</td></tr>
                    <tr><td><strong>GST Number:</strong></td><td>{quote_request.gst_number or 'Not provided'}</td></tr>
                </table>
            </div>
            <div class="col-md-6">
                <h6 class="text-primary">Product Information</h6>
                <table class="table table-sm">
                    <tr><td><strong>Product Type:</strong></td><td>{quote_request.product_type.prdt_desc}</td></tr>
                    <tr><td><strong>Quantity:</strong></td><td>{quote_request.quantity}</td></tr>
                    <tr><td><strong>Status:</strong></td><td>
                        <span class="badge bg-{'warning' if quote_request.status == 'new' else 'info' if quote_request.status == 'reviewed' else 'primary' if quote_request.status == 'quoted' else 'success' if quote_request.status == 'accepted' else 'danger' if quote_request.status == 'rejected' else 'secondary'}">
                            {quote_request.status.title()}
                        </span>
                    </td></tr>
                    <tr><td><strong>Created:</strong></td><td>{quote_request.created_at.strftime('%B %d, %Y at %I:%M %p')}</td></tr>
                </table>
            </div>
        </div>
        
        <div class="row mt-3">
            <div class="col-12">
                <h6 class="text-primary">Address Information</h6>
                <p>{quote_request.get_address_info()}</p>
            </div>
        </div>
        
        <div class="row mt-3">
            <div class="col-12">
                <h6 class="text-primary">Additional Requirements</h6>
                <p>{quote_request.additional_requirements or 'No additional requirements specified'}</p>
            </div>
        </div>
        
        <div class="row mt-3">
            <div class="col-12">
                <h6 class="text-primary">Email Status</h6>
                <table class="table table-sm">
                    <tr><td><strong>Email Sent:</strong></td><td>
                        <span class="badge bg-{'success' if quote_request.email_sent else 'danger'}">
                            <i class="fas fa-{'check' if quote_request.email_sent else 'times'}"></i>
                            {'Sent' if quote_request.email_sent else 'Not Sent'}
                        </span>
                    </td></tr>
                    <tr><td><strong>Sent At:</strong></td><td>{quote_request.email_sent_at.strftime('%B %d, %Y at %I:%M %p') if quote_request.email_sent_at else 'N/A'}</td></tr>
                </table>
            </div>
        </div>
        
        <div class="row mt-3">
            <div class="col-12">
                <h6 class="text-primary">Admin Notes</h6>
                <textarea class="form-control" rows="3" placeholder="Add admin notes here...">{quote_request.admin_notes or ''}</textarea>
            </div>
        </div>
        """
        
        return JsonResponse({
            'success': True,
            'html': html_content
        })
        
    except QuoteRequest.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Quote request not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

@admin_required
@csrf_exempt
def update_quote_request_status(request, request_id):
    """API endpoint to update quote request status"""
    from .models import QuoteRequest
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        new_status = data.get('status')
        
        if not new_status:
            return JsonResponse({'success': False, 'message': 'Status is required'}, status=400)
        
        quote_request = QuoteRequest.objects.get(id=request_id)
        quote_request.status = new_status
        quote_request.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Status updated to {new_status}'
        })
        
    except QuoteRequest.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Quote request not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

# ============================================================================
# PRODUCT MANAGEMENT VIEWS (New Structure)
# ============================================================================

@admin_required
def add_fusiontec_product(request):
    """Add new Fusiontec product"""
    from .models import ProductMasterV2
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('add_fusiontec_product')
        
        try:
            product = ProductMasterV2.objects.create(
                name=name,
                description=description,
                product_type='fusiontec'
            )
            
            if image:
                product.image = image
                product.save()
            
            messages.success(request, 'Fusiontec product added successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    return render(request, 'products/admin/add_fusiontec_product.html')

@admin_required
def edit_fusiontec_product(request, product_id):
    """Edit Fusiontec product"""
    from .models import ProductMasterV2
    product = get_object_or_404(ProductMasterV2, id=product_id, product_type='fusiontec')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('edit_fusiontec_product', product_id=product_id)
        
        try:
            product.name = name
            product.description = description
            
            if image:
                product.image = image
            
            product.save()
            messages.success(request, 'Fusiontec product updated successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error updating product: {str(e)}')
    
    context = {'product': product}
    return render(request, 'products/admin/edit_fusiontec_product.html', context)

@admin_required
def add_biz_product(request):
    """Add new Biz product"""
    from .models import ProductMasterV2
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('add_biz_product')
        
        try:
            product = ProductMasterV2.objects.create(
                name=name,
                description=description,
                product_type='biz'
            )
            
            if image:
                product.image = image
                product.save()
            
            messages.success(request, 'Biz product added successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    return render(request, 'products/admin/add_biz_product.html')

@admin_required
def edit_biz_product(request, product_id):
    """Edit Biz product"""
    from .models import ProductMasterV2
    product = get_object_or_404(ProductMasterV2, id=product_id, product_type='biz')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('edit_biz_product', product_id=product_id)
        
        try:
            product.name = name
            product.description = description
            
            if image:
                product.image = image
            
            product.save()
            messages.success(request, 'Biz product updated successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error updating product: {str(e)}')
    
    context = {'product': product}
    return render(request, 'products/admin/edit_biz_product.html', context)

@admin_required
def add_emudhra_product(request):
    """Add new e-Mudhra product"""
    from .models import ProductMasterV2
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('add_emudhra_product')
        
        try:
            product = ProductMasterV2.objects.create(
                name=name,
                description=description,
                product_type='emudhra'
            )
            
            if image:
                product.image = image
                product.save()
            
            messages.success(request, 'e-Mudhra product added successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    return render(request, 'products/admin/add_emudhra_product.html')

@admin_required
def edit_emudhra_product(request, product_id):
    """Edit e-Mudhra product"""
    from .models import ProductMasterV2
    product = get_object_or_404(ProductMasterV2, id=product_id, product_type='emudhra')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('edit_emudhra_product', product_id=product_id)
        
        try:
            product.name = name
            product.description = description
            
            if image:
                product.image = image
            
            product.save()
            messages.success(request, 'e-Mudhra product updated successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error updating product: {str(e)}')
    
    context = {'product': product}
    return render(request, 'products/admin/edit_emudhra_product.html', context)

@admin_required
def add_tally_product(request):
    """Add new Tally product"""
    from .models import ProductMasterV2
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('add_tally_product')
        
        try:
            product = ProductMasterV2.objects.create(
                name=name,
                description=description,
                product_type='tally'
            )
            
            if image:
                product.image = image
                product.save()
            
            messages.success(request, 'Tally product added successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    return render(request, 'products/admin/add_tally_product.html')

@admin_required
def edit_tally_product(request, product_id):
    """Edit Tally product"""
    from .models import ProductMasterV2
    product = get_object_or_404(ProductMasterV2, id=product_id, product_type='tally')
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        
        if not name:
            messages.error(request, 'Product name is required.')
            return redirect('edit_tally_product', product_id=product_id)
        
        try:
            product.name = name
            product.description = description
            
            if image:
                product.image = image
            
            product.save()
            messages.success(request, 'Tally product updated successfully.')
            return redirect('custom_admin_products')
        except Exception as e:
            messages.error(request, f'Error updating product: {str(e)}')
    
    context = {'product': product}
    return render(request, 'products/admin/edit_tally_product.html', context)

# ============================================================================
# SETTINGS MANAGEMENT VIEWS
# ============================================================================

@admin_required
def edit_razorpay_info(request, info_id):
    """Edit Razorpay Settings (PaymentSettings model)."""
    from .models import PaymentSettings
    razorpay = get_object_or_404(PaymentSettings, id=info_id, setting_type='razorpay')

    if request.method == 'POST':
        razorpay.title = request.POST.get('title', '').strip()
        razorpay.description = request.POST.get('description', '').strip()
        razorpay.payment_button_id = request.POST.get('payment_button_id', '').strip()
        # Optional keys if later added to the form
        razorpay.razorpay_key_id = request.POST.get('key_id', razorpay.razorpay_key_id or '').strip() or razorpay.razorpay_key_id
        razorpay.razorpay_key_secret = request.POST.get('key_secret', razorpay.razorpay_key_secret or '').strip() or razorpay.razorpay_key_secret
        razorpay.is_active = True
        razorpay.save()
        messages.success(request, 'Razorpay settings updated successfully.')
        return redirect('custom_admin_settings')

    context = { 'razorpay_info': razorpay }
    return render(request, 'products/admin/edit_razorpay_info.html', context)

@admin_required
def add_razorpay_info(request):
    """Add new Razorpay Settings (PaymentSettings model)."""
    from .models import PaymentSettings

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        payment_button_id = request.POST.get('payment_button_id', '').strip()

        if not title or not description or not payment_button_id:
            messages.error(request, 'Please fill Title, Description and Payment Button ID.')
            return redirect('add_razorpay_info')

        # Ensure only one active Razorpay settings row
        PaymentSettings.objects.filter(setting_type='razorpay').delete()
        PaymentSettings.objects.create(
            setting_type='razorpay',
            title=title,
            description=description,
            payment_button_id=payment_button_id,
            is_active=True,
        )
        messages.success(request, 'Razorpay settings added successfully.')
        return redirect('custom_admin_settings')

    return render(request, 'products/admin/add_razorpay_info.html')

@admin_required
def edit_qr_info(request, info_id):
    """Edit QR Code Settings (PaymentSettings model)."""
    from .models import PaymentSettings
    qr = get_object_or_404(PaymentSettings, id=info_id, setting_type='qr_code')

    if request.method == 'POST':
        company_name = request.POST.get('company_name', '').strip()
        upi_id = request.POST.get('upi_id', '').strip()
        qr_file = request.FILES.get('qr_image')

        if not all([company_name, upi_id]):
            messages.error(request, 'Company Name and UPI ID are required.')
            return redirect('edit_qr_info', info_id=qr.id)

        qr.title = company_name
        qr.upi_id = upi_id
        if qr_file:
            qr.qr_image = qr_file
        qr.is_active = True
        qr.save()
        messages.success(request, 'QR code settings updated successfully.')
        return redirect('custom_admin_settings')

    context = { 'qr_info': qr }
    return render(request, 'products/admin/edit_qr_info.html', context)

@admin_required
def add_qr_info(request):
    """Add QR Code Settings (PaymentSettings model)."""
    from .models import PaymentSettings

    if request.method == 'POST':
        company_name = request.POST.get('company_name', '').strip()
        upi_id = request.POST.get('upi_id', '').strip()
        qr_file = request.FILES.get('qr_image')

        if not all([company_name, upi_id]):
            messages.error(request, 'All fields are required.')
            return redirect('add_qr_info')

        # Keep only one active QR settings row
        PaymentSettings.objects.filter(setting_type='qr_code').delete()
        qr = PaymentSettings(
            setting_type='qr_code',
            title=company_name,
            upi_id=upi_id,
            is_active=True,
        )
        if qr_file:
            qr.qr_image = qr_file
        qr.save()
        messages.success(request, 'QR code settings added successfully.')
        return redirect('custom_admin_settings')

    return render(request, 'products/admin/add_qr_info.html')

@admin_required
def edit_bank_info(request, info_id):
    """Edit Bank Transfer Settings (PaymentSettings model)."""
    from .models import PaymentSettings
    bank = get_object_or_404(PaymentSettings, id=info_id, setting_type='bank_transfer')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        account_name = request.POST.get('account_name', '').strip()
        account_number = request.POST.get('account_number', '').strip()
        ifsc_code = request.POST.get('ifsc_code', '').strip()
        bank_name = request.POST.get('bank_name', '').strip()

        if not all([title, account_name, account_number, ifsc_code, bank_name]):
            messages.error(request, 'All fields are required.')
            return redirect('edit_bank_info', info_id=bank.id)

        bank.title = title
        bank.account_name = account_name
        bank.account_number = account_number
        bank.ifsc_code = ifsc_code
        bank.bank_name = bank_name
        bank.is_active = True
        bank.save()
        messages.success(request, 'Bank transfer settings updated successfully.')
        return redirect('custom_admin_settings')

    context = { 'bank_info': bank }
    return render(request, 'products/admin/edit_bank_info.html', context)

@admin_required
def add_bank_info(request):
    """Add Bank Transfer Settings (PaymentSettings model)."""
    from .models import PaymentSettings

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        account_name = request.POST.get('account_name', '').strip()
        account_number = request.POST.get('account_number', '').strip()
        ifsc_code = request.POST.get('ifsc_code', '').strip()
        bank_name = request.POST.get('bank_name', '').strip()

        if not all([title, account_name, account_number, ifsc_code, bank_name]):
            messages.error(request, 'All fields are required.')
            return redirect('add_bank_info')

        # Keep only one active bank transfer settings row
        PaymentSettings.objects.filter(setting_type='bank_transfer').delete()
        PaymentSettings.objects.create(
            setting_type='bank_transfer',
            title=title,
            account_name=account_name,
            account_number=account_number,
            ifsc_code=ifsc_code,
            bank_name=bank_name,
            is_active=True,
        )
        messages.success(request, 'Bank transfer settings added successfully.')
        return redirect('custom_admin_settings')

    return render(request, 'products/admin/add_bank_info.html')

# ============================================================================
# API ENDPOINTS
# ============================================================================

@csrf_exempt
def get_states(request):
    """Get list of states for forms"""
    import json
    import os
    from django.conf import settings
    
    # Load states from JSON file
    json_path = os.path.join(settings.BASE_DIR, 'data', 'states_districts.json')
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Extract state names from the JSON structure
        states = [item['state'] for item in data.get('states', [])]
        return JsonResponse({'states': states})
    except Exception as e:
        print(f"Error loading states: {e}")
        return JsonResponse({'states': []})

@csrf_exempt
def get_districts(request, state):
    """Get districts for a specific state"""
    import json
    import os
    from django.conf import settings
    
    # Load districts from JSON file
    json_path = os.path.join(settings.BASE_DIR, 'data', 'states_districts.json')
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Find the state and get its districts
        districts = []
        for item in data.get('states', []):
            if item['state'] == state:
                districts = item.get('districts', [])
                break
        
        return JsonResponse({'districts': districts})
    except Exception as e:
        print(f"Error loading districts: {e}")
        return JsonResponse({'districts': []})

@csrf_exempt
def get_submission_details(request, submission_id):
    """API endpoint to get submission details for modal"""
    try:
        submission = ProductFormSubmission.objects.get(id=submission_id)
        
        # Render the details HTML
        html_content = render_to_string('products/admin/submission_details.html', {
            'submission': submission
        })
        
        return JsonResponse({
            'success': True,
            'html': html_content
        })
        
    except ProductFormSubmission.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Submission not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
def convert_submission_to_quote(request, submission_id):
    """API endpoint to convert submission to quote"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid method'}, status=405)
    
    try:
        submission = ProductFormSubmission.objects.get(id=submission_id)
        
        if submission.status == 'converted':
            return JsonResponse({
                'success': False,
                'message': 'Submission already converted'
            })
        
        # Convert to quote
        quote = submission.convert_to_quote()
        
        return JsonResponse({
            'success': True,
            'message': 'Successfully converted to quote',
            'quote_id': quote.id
        })
        
    except ProductFormSubmission.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Submission not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error converting: {str(e)}'
        }, status=500)

# ============================================================================
# ERROR HANDLERS
# ============================================================================

def handler404(request, exception):
    """Custom 404 error handler"""
    return render(request, 'products/404.html', status=404)

def handler500(request):
    """Custom 500 error handler"""
    return render(request, 'products/500.html', status=500)

# ============================================================================
# DSC CHECKING ADMIN VIEW
# ============================================================================

@admin_required
def dsc_checking_admin(request):
    """DSC Checking admin panel with date filtering, pagination, and Excel export"""
    
    # Get filter parameters
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    per_page = request.GET.get('per_page', '10')
    export_excel = request.GET.get('export_excel', '')
    
    # Base queryset
    queryset = DscSubmission.objects.all().order_by('-created_at')
    
    # Debug: Show some sample dates from database
    sample_submissions = DscSubmission.objects.all()[:5]
    print(f"[DSC_CHECKING_DEBUG] Sample submission dates:")
    for i, sub in enumerate(sample_submissions):
        print(f"  {i+1}. ID: {sub.id}, Date: {sub.created_at}, Name: {sub.name}")
    
    # Apply date filters
    date_error = None
    
    if from_date:
        try:
            from_date_parsed = datetime.strptime(from_date, '%Y-%m-%d').date()
            # Convert to timezone-aware datetime for the start of the day
            from_datetime = timezone.make_aware(datetime.combine(from_date_parsed, datetime.min.time()))
            queryset = queryset.filter(created_at__gte=from_datetime)
            print(f"Applied from_date filter: {from_datetime}")
        except ValueError as e:
            date_error = f"Invalid from date format: {from_date}"
            print(f"From date parsing error: {e}")
    
    if to_date:
        try:
            from datetime import time
            to_date_parsed = datetime.strptime(to_date, '%Y-%m-%d').date()
            # Convert to timezone-aware datetime for the end of the day
            to_datetime = timezone.make_aware(datetime.combine(to_date_parsed, time(23, 59, 59, 999999)))
            queryset = queryset.filter(created_at__lte=to_datetime)
            print(f"Applied to_date filter: {to_datetime}")
        except ValueError as e:
            date_error = f"Invalid to date format: {to_date}"
            print(f"To date parsing error: {e}")
    
    # Validate date range
    if from_date and to_date and not date_error:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            if from_date_obj > to_date_obj:
                date_error = "From Date cannot be later than To Date."
        except ValueError:
            pass
    
    # Handle Excel export
    if export_excel == 'true':
        return export_dsc_data_to_excel(queryset, from_date, to_date)
    
    # Pagination
    try:
        per_page_int = int(per_page)
        if per_page_int not in [10, 20, 50, 100]:
            per_page_int = 10
    except:
        per_page_int = 10
    
    paginator = Paginator(queryset, per_page_int)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    total_submissions = queryset.count()
    new_submissions = queryset.filter(status='submitted').count()
    payment_received = queryset.filter(status='payment_received').count()
    processing = queryset.filter(status='processing').count()
    completed = queryset.filter(status='completed').count()
    
    # Debug information
    print(f"[DSC_CHECKING_DEBUG] from_date: {from_date}, to_date: {to_date}")
    print(f"[DSC_CHECKING_DEBUG] Total records after filtering: {total_submissions}")
    print(f"[DSC_CHECKING_DEBUG] Date error: {date_error}")
    
    # Get all submission dates for debugging
    if total_submissions > 0:
        first_submission = queryset.first()
        last_submission = queryset.last()
        print(f"[DSC_CHECKING_DEBUG] First submission date: {first_submission.created_at}")
        print(f"[DSC_CHECKING_DEBUG] Last submission date: {last_submission.created_at}")
    
    context = {
        'page_obj': page_obj,
        'from_date': from_date,
        'to_date': to_date,
        'per_page': per_page,
        'total_submissions': total_submissions,
        'new_submissions': new_submissions,
        'payment_received': payment_received,
        'processing': processing,
        'completed': completed,
        'per_page_options': [10, 20, 50, 100],
        'date_error': date_error,
    }
    
    return render(request, 'products/admin/dsc_checking.html', context)

def export_dsc_data_to_excel(queryset, from_date, to_date):
    """Export DSC data to Excel file"""
    from django.http import HttpResponse
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DSC Data"
    
    # Define headers as per requirement
    headers = [
        'S.NO', 'Date', 'Application No.', 'CUSTOMER NAME', 'MOBILE NUMBER', 
        'E-MAIL', 'NEW / RENEWAL', 'REFERENCE NAME', 'REFERENCE MOBILE', 
        'Company Name / Individual Name', 'TYPE', 'E-SIGN / pantan e-sing', 
        'GST', 'AMOUNT', 'TOKEN', 'ADDRESS', 'Order No', 'UPI ID FUSIONTEC'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add data rows
    for row_num, submission in enumerate(queryset, 2):
        # S.NO
        ws.cell(row=row_num, column=1, value=row_num - 1)
        
        # Date
        ws.cell(row=row_num, column=2, value=submission.created_at.strftime('%d/%m/%Y'))
        
        # Application No.
        ws.cell(row=row_num, column=3, value=f"DSC-{submission.id}")
        
        # CUSTOMER NAME
        ws.cell(row=row_num, column=4, value=submission.name)
        
        # MOBILE NUMBER
        ws.cell(row=row_num, column=5, value=submission.mobile)
        
        # E-MAIL
        ws.cell(row=row_num, column=6, value=submission.email)
        
        # NEW / RENEWAL (assuming all are NEW for DSC)
        ws.cell(row=row_num, column=7, value="NEW")
        
        # REFERENCE NAME
        ws.cell(row=row_num, column=8, value=submission.reference_name or "")
        
        # REFERENCE MOBILE
        ws.cell(row=row_num, column=9, value=submission.reference_contact or "")
        
        # Company Name / Individual Name
        company_or_individual = submission.company_name if submission.company_name else "Individual"
        ws.cell(row=row_num, column=10, value=company_or_individual)
        
        # TYPE
        dsc_type = f"{submission.class_type} - {submission.user_type} - {submission.cert_type} - {submission.validity}"
        ws.cell(row=row_num, column=11, value=dsc_type)
        
        # E-SIGN / pantan e-sing
        ws.cell(row=row_num, column=12, value=submission.cert_type.upper())
        
        # GST
        ws.cell(row=row_num, column=13, value=submission.gst_number or "")
        
        # AMOUNT
        ws.cell(row=row_num, column=14, value=float(submission.quoted_price))
        
        # TOKEN
        token_status = "Yes" if submission.include_token else "No"
        ws.cell(row=row_num, column=15, value=token_status)
        
        # ADDRESS
        ws.cell(row=row_num, column=16, value=submission.address or "")
        
        # Order No
        ws.cell(row=row_num, column=17, value=submission.razorpay_order_id or "")
        
        # UPI ID FUSIONTEC (placeholder)
        ws.cell(row=row_num, column=18, value="dsc@fusiontec.com")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create filename with date range
    if from_date and to_date:
        filename = f"DSC_Data_{from_date}_to_{to_date}.xlsx"
    elif from_date:
        filename = f"DSC_Data_from_{from_date}.xlsx"
    elif to_date:
        filename = f"DSC_Data_to_{to_date}.xlsx"
    else:
        filename = f"DSC_Data_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response
